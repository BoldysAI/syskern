"""Excel export for a simulation (CDC §6.9 — partage offline).

Builds a 3-sheet workbook from a `Simulation` and its lines:
  - "Synthèse"          : market params, chain summary, mix/margins, aggregates
  - "Résultats"         : one row per SKU with the UI table columns
  - "Breakdown détaillé": per line, the PA/PV chain step by step

Pure presentation — no pricing logic happens here (the engine already froze
every result on the lines).
"""

from __future__ import annotations

import io
from decimal import Decimal, InvalidOperation
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from .models import Simulation, SimulationLine

_HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_SECTION_FONT = Font(bold=True, size=11, color="1F3864")
_ALT_FILL = PatternFill(start_color="EEF2F7", end_color="EEF2F7", fill_type="solid")
_EUR_FORMAT = "#,##0.00 €"

_RESULT_COLUMNS: list[tuple[str, int]] = [
    ("SKU", 20),
    ("Nom", 40),
    ("Gamme", 20),
    ("Stock", 12),
    ("PAMP réel (EUR)", 16),
    ("PA net (EUR)", 14),
    ("PAMP prévisionnel (EUR)", 20),
    ("PR (EUR)", 14),
    ("Marge effective (%)", 16),
    ("Mix effectif (%)", 14),
    ("PV (EUR)", 14),
    ("Statut", 14),
]

_STATUS_LABELS = {
    "ok": "OK",
    "pending": "En attente",
    "warning": "Avertissement",
    "error": "Erreur",
    "dirty": "Modifié",
}


def _num(value) -> float | None:
    """Coerce a Decimal/str money value to float for Excel numeric cells.

    Display-only conversion (the canonical Decimal stays in the DB / engine).
    """
    if value is None or value == "":
        return None
    try:
        return float(Decimal(str(value)))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _style_header_row(ws, headers: list[tuple[str, int]]) -> None:
    for col_idx, (label, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"


def _aggregate(lines: list[SimulationLine]) -> dict:
    pv = [v for v in (_num(ln.pv_eur) for ln in lines) if v is not None]
    pa = [v for v in (_num(ln.pa_net_eur) for ln in lines) if v is not None]
    pr = [v for v in (_num(ln.pr_eur) for ln in lines) if v is not None]
    return {
        "count": len(lines),
        "avg_pa": sum(pa) / len(pa) if pa else None,
        "avg_pr": sum(pr) / len(pr) if pr else None,
        "avg_pv": sum(pv) / len(pv) if pv else None,
        "min_pv": min(pv) if pv else None,
        "max_pv": max(pv) if pv else None,
        "warnings": sum(1 for ln in lines if ln.status == "warning"),
        "errors": sum(1 for ln in lines if ln.status == "error"),
    }


def _build_synthese(ws, simulation: Simulation, lines: list[SimulationLine]) -> None:
    ws.title = "Synthèse"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 32
    row = 1

    def section(title: str) -> None:
        nonlocal row
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = _SECTION_FONT
        row += 1

    def kv(label: str, value, *, eur: bool = False) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        if eur and isinstance(value, (int, float)):
            cell.number_format = _EUR_FORMAT
        row += 1

    section("Simulation")
    kv("Libellé", simulation.label)
    kv("Type", simulation.get_simulation_type_display())
    kv("Statut", simulation.get_status_display())
    kv("Projet", simulation.project_name or "—")
    kv(
        "Dernier calcul",
        simulation.last_calculated_at.strftime("%Y-%m-%d %H:%M")
        if simulation.last_calculated_at
        else "—",
    )
    kv(
        "Snapshot Odoo",
        simulation.odoo_snapshot_at.strftime("%Y-%m-%d %H:%M")
        if simulation.odoo_snapshot_at
        else "—",
    )
    row += 1

    section("Paramètres globaux")
    kv("Mix stock/achat (%)", simulation.stock_purchase_mix_pct)
    kv("Marge Symea (%)", (_num(simulation.symea_margin_rate) or 0) * 100)
    kv("Marge Syskern (%)", (_num(simulation.syskern_margin_rate) or 0) * 100)
    row += 1

    section("Paramètres marché")
    for key, value in (simulation.market_params or {}).items():
        kv(str(key), str(value))
    if not (simulation.market_params or {}):
        kv("—", "Aucun paramètre marché figé")
    row += 1

    agg = _aggregate(lines)
    section("Agrégats")
    kv("Nombre de lignes", agg["count"])
    kv("PA net moyen", agg["avg_pa"], eur=True)
    kv("PR moyen", agg["avg_pr"], eur=True)
    kv("PV moyen", agg["avg_pv"], eur=True)
    kv("PV min", agg["min_pv"], eur=True)
    kv("PV max", agg["max_pv"], eur=True)
    kv("Avertissements", agg["warnings"])
    kv("Erreurs", agg["errors"])


def _build_resultats(ws, lines: list[SimulationLine]) -> None:
    ws.title = "Résultats"
    _style_header_row(ws, _RESULT_COLUMNS)
    eur_cols = {5, 6, 7, 8, 11}  # 1-based column indexes that hold EUR amounts

    for row_idx, line in enumerate(lines, start=2):
        product = line.product
        margin = _num(line.effective_margin_rate)
        values = [
            product.sku_code,
            product.designation,
            product.range,
            _num(product.stock_quantity),
            _num(product.pamp_eur),
            _num(line.pa_net_eur),
            _num(line.pamp_predictive_eur),
            _num(line.pr_eur),
            round(margin * 100, 2) if margin is not None else None,
            line.effective_mix_pct,
            _num(line.pv_eur),
            _STATUS_LABELS.get(line.status, line.status),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in eur_cols and isinstance(value, (int, float)):
                cell.number_format = _EUR_FORMAT
            if row_idx % 2 == 0:
                cell.fill = _ALT_FILL
            cell.alignment = Alignment(vertical="center")


def _build_breakdown(ws, lines: list[SimulationLine]) -> None:
    ws.title = "Breakdown détaillé"
    headers = [
        ("SKU", 20),
        ("Chaîne", 12),
        ("Ordre", 8),
        ("Module", 24),
        ("Appliqué", 10),
        ("Entrée", 16),
        ("Devise", 10),
        ("Sortie", 16),
        ("Devise", 10),
    ]
    _style_header_row(ws, headers)
    row = 2

    for line in lines:
        breakdown = line.calculation_breakdown or {}
        if "error" in breakdown:
            ws.cell(row=row, column=1, value=line.product.sku_code)
            ws.cell(row=row, column=4, value=f"ERREUR : {breakdown['error']}")
            row += 1
            continue
        for chain_key, chain_label in (("purchase", "PA"), ("sale", "PV")):
            chain = breakdown.get(chain_key) or {}
            for step in chain.get("steps", []):
                inp = step.get("input_price", {})
                out = step.get("output_price", {})
                ws.cell(row=row, column=1, value=line.product.sku_code)
                ws.cell(row=row, column=2, value=chain_label)
                ws.cell(row=row, column=3, value=step.get("order"))
                ws.cell(row=row, column=4, value=step.get("module"))
                ws.cell(row=row, column=5, value="Oui" if step.get("applied") else "Non")
                c_in = ws.cell(row=row, column=6, value=_num(inp.get("amount")))
                c_in.number_format = "#,##0.0000"
                ws.cell(row=row, column=7, value=inp.get("currency"))
                c_out = ws.cell(row=row, column=8, value=_num(out.get("amount")))
                c_out.number_format = "#,##0.0000"
                ws.cell(row=row, column=9, value=out.get("currency"))
                row += 1


def build_simulation_xlsx(simulation: Simulation) -> bytes:
    """Build the multi-sheet Excel workbook for *simulation* and return bytes.

    The simulation should be loaded with `lines__product` prefetched.
    """
    lines = list(simulation.lines.select_related("product").all())
    wb = openpyxl.Workbook()

    _build_synthese(wb.active, simulation, lines)
    _build_resultats(wb.create_sheet(), lines)
    _build_breakdown(wb.create_sheet(), lines)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
