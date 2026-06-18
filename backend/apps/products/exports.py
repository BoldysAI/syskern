"""Excel export for the product catalog (CDC §4.1.1).

Generates a .xlsx workbook from an arbitrary Product queryset, with an
optional caller-selected subset of columns.
"""

from __future__ import annotations

import io
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from django.db.models import QuerySet

    from .models import Product

# ─── Column registry ───────────────────────────────────────────────────────────
# key → (header label, attribute accessor, column width). The accessor
# `__active_supplier__` is resolved specially. Keep this list in sync with the
# frontend column picker (frontend/src/app/catalog/page.tsx).
_COLUMN_REGISTRY: dict[str, tuple[str, str, int]] = {
    "sku_code": ("SKU", "sku_code", 20),
    "name": ("Nom", "name", 40),
    "universe": ("Univers", "universe", 20),
    "family": ("Famille", "family", 20),
    "range": ("Gamme", "range", 20),
    "sub_range": ("Sous-gamme", "sub_range", 20),
    "brand": ("Marque", "brand", 16),
    "active_supplier": ("Fournisseur actif", "__active_supplier__", 24),
    "factory_code": ("Code usine", "factory_code", 14),
    "stock_quantity": ("Stock (unités)", "stock_quantity", 16),
    "pamp_eur": ("PAMP (EUR)", "pamp_eur", 14),
    "is_copper_indexed": ("Indexé cuivre", "is_copper_indexed", 14),
    "is_active": ("Actif", "is_active", 10),
}

# Ordered default column set when the caller does not specify `columns`.
DEFAULT_COLUMNS: list[str] = [
    "sku_code",
    "name",
    "universe",
    "family",
    "range",
    "sub_range",
    "brand",
    "active_supplier",
    "stock_quantity",
    "pamp_eur",
    "is_copper_indexed",
    "is_active",
]

_BOOL_COLUMNS = {"is_copper_indexed", "is_active"}

_HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_ALT_FILL = PatternFill(start_color="EEF2F7", end_color="EEF2F7", fill_type="solid")


def _resolve_columns(columns: list[str] | None) -> list[str]:
    """Validate caller-selected columns, preserving order; fall back to default."""
    if not columns:
        return DEFAULT_COLUMNS
    selected = [c for c in columns if c in _COLUMN_REGISTRY]
    return selected or DEFAULT_COLUMNS


def _active_supplier_name(product: Product) -> str:
    for s in product.suppliers.all():
        if s.is_active:
            return s.supplier_name
    return ""


def build_products_xlsx(queryset: QuerySet[Product], columns: list[str] | None = None) -> bytes:
    """Build an Excel workbook from *queryset* and return the raw bytes.

    The queryset should already have `prefetch_related("suppliers")` applied
    for efficient supplier name resolution. *columns* is an optional ordered
    list of column keys (see `_COLUMN_REGISTRY`); defaults to `DEFAULT_COLUMNS`.
    """
    keys = _resolve_columns(columns)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catalogue produits"

    for col_idx, key in enumerate(keys, start=1):
        label, _, width = _COLUMN_REGISTRY[key]
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for row_idx, product in enumerate(queryset, start=2):
        for col_idx, key in enumerate(keys, start=1):
            _, attr, _ = _COLUMN_REGISTRY[key]
            if attr == "__active_supplier__":
                value = _active_supplier_name(product)
            elif key in _BOOL_COLUMNS:
                value = "Oui" if getattr(product, attr) else "Non"
            else:
                raw = getattr(product, attr, None)
                value = str(raw) if raw is not None else ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx % 2 == 0:
                cell.fill = _ALT_FILL
            cell.alignment = Alignment(vertical="center")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
