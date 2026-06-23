"""Tariff-offer Excel generation (CDC §7.2.3).

One workbook per client: a single "Tarifs" sheet with a branded header, the
user-selected columns (ordered), and a conditions footer. Column headers are
translated FR/EN/ES from an in-code dictionary (CDC §10.5.4); prices are already
converted to the offer currency when the offer lines are built, so this module
only formats them.
"""

from __future__ import annotations

import io
from collections.abc import Callable
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Column registry ──────────────────────────────────────────────────────────
# Each column: multilingual header + a value extractor (line, offer, client).
# Extractors return plain primitives (str / int / float / None) — never Decimal,
# which openpyxl cannot serialise.

_Number = float  # cells we want Excel to treat as numbers


def _price(line, offer, client) -> _Number | None:
    return float(round(line.final_price, 2)) if line.final_price is not None else None


def _product_attr(name: str) -> Callable:
    return lambda line, offer, client: getattr(line.product, name, None) or None


def _weight(line, offer, client) -> _Number | None:
    w = line.product.unit_weight_kg
    return float(w) if w is not None else None


_COLUMN_REGISTRY: dict[str, dict[str, Any]] = {
    "sku_code": {"h": {"fr": "Réf. SKU", "en": "SKU", "es": "SKU"}, "v": _product_attr("sku_code")},
    "name": {
        "h": {"fr": "Désignation", "en": "Name", "es": "Designación"},
        "v": _product_attr("name"),
    },
    "brand": {"h": {"fr": "Marque", "en": "Brand", "es": "Marca"}, "v": _product_attr("brand")},
    "universe": {
        "h": {"fr": "Univers", "en": "Universe", "es": "Universo"},
        "v": _product_attr("universe"),
    },
    "family": {
        "h": {"fr": "Famille", "en": "Family", "es": "Familia"},
        "v": _product_attr("family"),
    },
    "range": {"h": {"fr": "Gamme", "en": "Range", "es": "Gama"}, "v": _product_attr("range")},
    "sub_range": {
        "h": {"fr": "Sous-gamme", "en": "Sub-range", "es": "Subgama"},
        "v": _product_attr("sub_range"),
    },
    "unit_price": {
        "h": {"fr": "Prix unitaire", "en": "Unit price", "es": "Precio unitario"},
        "v": _price,
        "number_format": "#,##0.00",
    },
    "currency": {
        "h": {"fr": "Devise", "en": "Currency", "es": "Divisa"},
        "v": lambda line, offer, client: offer.currency,
    },
    "incoterm": {
        "h": {"fr": "Incoterm", "en": "Incoterm", "es": "Incoterm"},
        "v": lambda line, offer, client: offer.incoterm,
    },
    "primary_packaging_qty": {
        "h": {"fr": "Cond. primaire", "en": "Primary pack", "es": "Emb. primario"},
        "v": _product_attr("primary_packaging_qty"),
    },
    "secondary_packaging_qty": {
        "h": {"fr": "Cond. secondaire", "en": "Secondary pack", "es": "Emb. secundario"},
        "v": _product_attr("secondary_packaging_qty"),
    },
    "tertiary_packaging_qty": {
        "h": {"fr": "Cond. tertiaire", "en": "Tertiary pack", "es": "Emb. terciario"},
        "v": _product_attr("tertiary_packaging_qty"),
    },
    "pallet_qty": {
        "h": {"fr": "Qté/palette", "en": "Qty/pallet", "es": "Cant./palé"},
        "v": _product_attr("pallet_qty"),
    },
    "unit_weight_kg": {
        "h": {"fr": "Poids unitaire (kg)", "en": "Unit weight (kg)", "es": "Peso unitario (kg)"},
        "v": _weight,
        "number_format": "#,##0.000",
    },
    "hs_code": {
        "h": {"fr": "Code SH", "en": "HS code", "es": "Código SA"},
        "v": _product_attr("hs_code"),
    },
    "gtin": {"h": {"fr": "GTIN", "en": "GTIN", "es": "GTIN"}, "v": _product_attr("gtin")},
    "client_name": {
        "h": {"fr": "Client", "en": "Client", "es": "Cliente"},
        "v": lambda line, offer, client: client.name if client else None,
    },
}

# Sensible default order when the wizard sends none.
DEFAULT_COLUMNS = [
    "sku_code",
    "name",
    "range",
    "unit_price",
    "currency",
    "primary_packaging_qty",
    "pallet_qty",
]

# Labels / static copy by language.
_SHEET_TITLE = {"fr": "Tarifs", "en": "Tariffs", "es": "Tarifas"}
_TAGLINE = {
    "fr": "Grille tarifaire",
    "en": "Price list",
    "es": "Lista de precios",
}
_LABELS = {
    "client": {"fr": "Client", "en": "Client", "es": "Cliente"},
    "offer": {"fr": "Offre", "en": "Offer", "es": "Oferta"},
    "validity": {"fr": "Validité", "en": "Validity", "es": "Validez"},
    "currency": {"fr": "Devise", "en": "Currency", "es": "Divisa"},
    "conditions": {
        "fr": "Conditions générales",
        "en": "General conditions",
        "es": "Condiciones generales",
    },
}
_CONDITIONS_TEXT = {
    "fr": "Prix indicatifs hors taxes, sous réserve de disponibilité. "
    "Les conditions générales de vente Syskern s'appliquent.",
    "en": "Indicative prices excluding tax, subject to availability. "
    "Syskern general terms of sale apply.",
    "es": "Precios indicativos sin impuestos, sujetos a disponibilidad. "
    "Se aplican las condiciones generales de venta de Syskern.",
}

_NAVY = "0F2137"
_ORANGE = "E07200"


def validate_columns(columns: list[str]) -> list[str]:
    """Raise ValueError on any unknown column key; return the list unchanged."""
    unknown = [c for c in columns if c not in _COLUMN_REGISTRY]
    if unknown:
        raise ValueError(f"Unknown column(s): {unknown}. Valid: {sorted(_COLUMN_REGISTRY)}")
    return columns


def available_columns(lang: str = "fr") -> list[dict]:
    """The column catalogue for the wizard ([{key, label}], translated)."""
    return [
        {"key": k, "label": v["h"].get(lang, v["h"]["fr"])} for k, v in _COLUMN_REGISTRY.items()
    ]


def build_tariff_xlsx(
    *,
    offer,
    client,
    lines: list,
    columns: list[str] | None,
    lang: str = "fr",
    fx_note: str = "",
) -> bytes:
    """Render one tariff workbook. ``lines`` must have ``product`` preloaded."""
    lang = lang if lang in {"fr", "en", "es"} else "fr"
    cols = columns or DEFAULT_COLUMNS

    wb = Workbook()
    ws = wb.active
    ws.title = _SHEET_TITLE[lang]

    # ── Branded header ──
    ws["A1"] = "SYSKERN"
    ws["A1"].font = Font(bold=True, size=20, color=_NAVY)
    ws["A2"] = _TAGLINE[lang]
    ws["A2"].font = Font(size=11, color=_ORANGE)

    row = 4
    meta = [
        (_LABELS["client"][lang], client.name if client else "—"),
        (_LABELS["offer"][lang], offer.label),
        (
            _LABELS["validity"][lang],
            " → ".join(filter(None, [str(offer.valid_from or ""), str(offer.valid_to or "")]))
            or "—",
        ),
        (_LABELS["currency"][lang], offer.currency),
    ]
    for label, value in meta:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1
    if fx_note:
        ws.cell(row=row, column=1, value=fx_note).font = Font(italic=True, size=9, color="666666")
        row += 1

    # ── Column header row ──
    header_row = row + 1
    for col_idx, key in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=_COLUMN_REGISTRY[key]["h"][lang])
        cell.fill = PatternFill("solid", fgColor=_NAVY)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # ── Data rows ──
    r = header_row + 1
    for line in lines:
        for col_idx, key in enumerate(cols, start=1):
            spec = _COLUMN_REGISTRY[key]
            cell = ws.cell(row=r, column=col_idx, value=spec["v"](line, offer, client))
            if spec.get("number_format"):
                cell.number_format = spec["number_format"]
        r += 1

    # ── Conditions footer ──
    r += 1
    ws.cell(row=r, column=1, value=_LABELS["conditions"][lang]).font = Font(bold=True)
    ws.cell(row=r + 1, column=1, value=_CONDITIONS_TEXT[lang]).font = Font(size=9, color="666666")

    # Column widths.
    for col_idx, key in enumerate(cols, start=1):
        header_len = len(_COLUMN_REGISTRY[key]["h"][lang])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(header_len + 4, 14), 48)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def fx_note_for(target_currency: str, market_params: dict, rate: Decimal) -> str:
    """Footnote tracing the applied conversion (CDC §7.2.5)."""
    if target_currency.upper() == "EUR":
        return ""
    return f"Taux de conversion appliqué : 1 EUR = {rate} {target_currency.upper()}"
