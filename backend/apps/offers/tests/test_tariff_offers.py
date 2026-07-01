"""Tests for tariff-offer Excel generation (CDC §7.2).

Covers the acceptance criteria: multi-client generation (N clients → N offers +
N files), column order, currency conversion, 400s for wrong simulation
state/type, valid Excel, and the download endpoint.
"""

from __future__ import annotations

from decimal import Decimal

import pytest
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.clients.models import Client
from apps.offers.models import GenerationStatus, Offer, OfferStatus, OfferType
from apps.offers.tasks import generate_tariff_offers_task, offer_export_path
from apps.products.models import Product
from apps.simulations.models import Simulation, SimulationLine

pytestmark = pytest.mark.django_db


@pytest.fixture()
def client_api() -> APIClient:
    return APIClient()


@pytest.fixture()
def export_dir(tmp_path, monkeypatch):
    """Redirect generated files to a tmp dir (shared by task + download view)."""
    monkeypatch.setattr("apps.offers.tasks.EXPORT_DIR", tmp_path)
    return tmp_path


@pytest.fixture()
def finalized_tariff():
    # Build as draft, add the lines, then finalize: the DB guard trigger
    # (simulation_lines_guard_finalized_parent) blocks line inserts once the
    # parent is finalized, so the lines must exist first.
    sim = Simulation.objects.create(
        label="Tarif Q3 2026",
        simulation_type="tariff",
        status="draft",
        market_params={"fx_eur_usd": "1.15", "fx_eur_rmb": "7.95"},
    )
    p1 = Product.objects.create(
        sku_code="KCFF6A4PZHDBL5-21",
        name="Câble cat7",
        range="Catégorie 7",
        primary_packaging_qty=10,
        pallet_qty=9,
    )
    p2 = Product.objects.create(
        sku_code="RACK-42U",
        name="Baie 42U",
        range="Racks",
        primary_packaging_qty=1,
        pallet_qty=4,
    )
    SimulationLine.objects.create(simulation=sim, product=p1, pv_eur=Decimal("487.70"), status="ok")
    SimulationLine.objects.create(
        simulation=sim, product=p2, pv_eur=Decimal("1200.00"), status="ok"
    )
    Simulation.objects.filter(pk=sim.pk).update(status="finalized")
    sim.refresh_from_db()
    return sim


@pytest.fixture()
def two_clients():
    return [
        Client.objects.create(name="Distributeur A"),
        Client.objects.create(name="Intégrateur B"),
    ]


def _params(clients, **over):
    p = {
        "client_ids": [str(c.id) for c in clients],
        "columns": [],
        "target_currency": "EUR",
        "language": "fr",
        "expiration_date": "2026-09-30",
        "incoterm": "EXW",
        "label": "Tarif Q3 2026",
    }
    p.update(over)
    return p


# ── Endpoint (202 / 400) ─────────────────────────────────────────────────────


def test_endpoint_returns_202(client_api, finalized_tariff, two_clients):
    resp = client_api.post(
        f"/api/simulations/{finalized_tariff.id}/generate-tariff-offers/",
        {"client_ids": [str(c.id) for c in two_clients], "expiration_date": "2026-09-30"},
        format="json",
    )
    assert resp.status_code == 202
    assert "task_id" in resp.json()
    assert resp.json()["client_count"] == 2


def test_draft_simulation_returns_400(client_api, two_clients):
    sim = Simulation.objects.create(label="Brouillon", simulation_type="tariff", status="draft")
    resp = client_api.post(
        f"/api/simulations/{sim.id}/generate-tariff-offers/",
        {"client_ids": [str(two_clients[0].id)]},
        format="json",
    )
    assert resp.status_code == 400


def test_project_simulation_returns_400(client_api, two_clients):
    sim = Simulation.objects.create(label="Projet", simulation_type="project", status="finalized")
    resp = client_api.post(
        f"/api/simulations/{sim.id}/generate-tariff-offers/",
        {"client_ids": [str(two_clients[0].id)]},
        format="json",
    )
    assert resp.status_code == 400


def test_unknown_column_rejected(client_api, finalized_tariff, two_clients):
    resp = client_api.post(
        f"/api/simulations/{finalized_tariff.id}/generate-tariff-offers/",
        {"client_ids": [str(two_clients[0].id)], "columns": ["bogus_column"]},
        format="json",
    )
    assert resp.status_code == 400


# ── Generation logic (run task eagerly) ──────────────────────────────────────


def _run(sim, params):
    return generate_tariff_offers_task.apply(args=[str(sim.id), params]).get()


def test_generates_one_offer_and_file_per_client(export_dir, finalized_tariff, two_clients):
    result = _run(finalized_tariff, _params(two_clients))

    assert result["count"] == 2
    assert Offer.objects.filter(offer_type=OfferType.TARIFF).count() == 2
    for entry in result["offers"]:
        offer = Offer.objects.get(id=entry["offer_id"])
        assert offer.status == OfferStatus.DRAFT
        # Excel is produced synchronously → generation is terminal (READY), not
        # left at the default PENDING (which would keep the UI polling — B1).
        assert offer.generation_status == GenerationStatus.READY
        assert len(offer.client_ids) == 1
        assert offer.lines.count() == 2  # both priced lines
        assert offer.generated_file_url == f"/api/offers/{offer.id}/download/"
        assert offer_export_path(offer.id).is_file()


def test_currency_conversion_eur_to_usd(export_dir, finalized_tariff, two_clients):
    _run(finalized_tariff, _params([two_clients[0]], target_currency="USD"))
    offer = Offer.objects.get(currency="USD")
    line = offer.lines.get(product__sku_code="KCFF6A4PZHDBL5-21")
    # 487.70 EUR × 1.15 = 560.8550 USD
    assert line.final_price == Decimal("560.8550")


def test_eur_keeps_price_unchanged(export_dir, finalized_tariff, two_clients):
    _run(finalized_tariff, _params([two_clients[0]], target_currency="EUR"))
    offer = Offer.objects.get(currency="EUR")
    line = offer.lines.get(product__sku_code="RACK-42U")
    assert line.final_price == Decimal("1200.0000")


def test_excel_opens_with_columns_in_order(export_dir, finalized_tariff, two_clients):
    cols = ["name", "sku_code", "unit_price", "currency"]
    _run(finalized_tariff, _params([two_clients[0]], columns=cols))
    offer = Offer.objects.first()

    wb = load_workbook(offer_export_path(offer.id))
    ws = wb["Tarifs"]
    expected = ["Désignation", "Réf. SKU", "Prix unitaire", "Devise"]
    # Find the header row by matching the first expected header in column 1.
    header_row = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == expected[0]
    )
    actual = [ws.cell(row=header_row, column=c).value for c in range(1, len(expected) + 1)]
    assert actual == expected
    # The next row holds data (the priced SKU).
    assert ws.cell(row=header_row + 1, column=2).value == "KCFF6A4PZHDBL5-21"


def test_english_headers(export_dir, finalized_tariff, two_clients):
    _run(
        finalized_tariff,
        _params([two_clients[0]], language="en", columns=["sku_code", "unit_price"]),
    )
    offer = Offer.objects.first()
    ws = load_workbook(offer_export_path(offer.id))["Tariffs"]
    header_row = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "SKU"
    )
    assert ws.cell(row=header_row, column=2).value == "Unit price"


# ── Download endpoint ────────────────────────────────────────────────────────


def test_download_endpoint(client_api, export_dir, finalized_tariff, two_clients):
    _run(finalized_tariff, _params([two_clients[0]]))
    offer = Offer.objects.first()
    resp = client_api.get(f"/api/offers/{offer.id}/download/")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp["Content-Type"]


def test_download_404_when_not_generated(client_api, finalized_tariff, two_clients):
    # An offer with no generated file → 404.
    offer = Offer.objects.create(
        simulation=finalized_tariff,
        offer_type=OfferType.TARIFF,
        label="x",
        client_ids=[str(two_clients[0].id)],
        currency="EUR",
        incoterm="EXW",
        export_format="excel",
    )
    resp = client_api.get(f"/api/offers/{offer.id}/download/")
    assert resp.status_code == 404
