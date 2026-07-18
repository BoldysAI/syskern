"""Tariff export — selectable market-parameter columns (FEEDBACK 1, CDC §7.2.3)."""

from __future__ import annotations

import io
from decimal import Decimal
from types import SimpleNamespace

from openpyxl import load_workbook

from apps.offers.services.excel import available_columns, build_tariff_xlsx, validate_columns

_MARKET_KEYS = {
    "copper_market",
    "copper_base_price",
    "copper_current_price",
    "fx_eur_rmb",
    "fx_eur_usd",
}


def test_market_param_columns_are_in_the_catalogue():
    keys = {c["key"] for c in available_columns("fr")}
    assert keys >= _MARKET_KEYS
    # And they validate (so the wizard can select them).
    assert validate_columns(sorted(_MARKET_KEYS)) == sorted(_MARKET_KEYS)


def test_build_tariff_xlsx_renders_market_params_from_simulation_snapshot():
    offer = SimpleNamespace(
        language="fr",
        currency="EUR",
        incoterm="EXW",
        label="Offre Test",
        valid_from=None,
        valid_to=None,
        simulation=SimpleNamespace(
            market_params={
                "copper_market": "LME",
                "copper_base_price": "9500",
                "fx_eur_rmb": "8.19",
            }
        ),
    )
    line = SimpleNamespace(
        final_price=Decimal("12.50"),
        product=SimpleNamespace(sku_code="P1"),
    )
    client = SimpleNamespace(name="ACME")
    cols = validate_columns(["sku_code", "copper_market", "copper_base_price", "fx_eur_rmb"])

    data = build_tariff_xlsx(offer=offer, client=client, lines=[line], columns=cols, lang="fr")
    ws = load_workbook(io.BytesIO(data)).active

    # Locate the header row, then read the single data row under it.
    header_row = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Réf. SKU"
    )
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, len(cols) + 1)]
    values = [ws.cell(row=header_row + 1, column=c).value for c in range(1, len(cols) + 1)]
    row = dict(zip(headers, values, strict=True))

    assert row["Marché cuivre"] == "LME"  # text
    assert row["Base cuivre"] == 9500.0  # numeric
    assert row["FX EUR→RMB"] == 8.19


def test_market_param_blank_when_snapshot_missing_key():
    offer = SimpleNamespace(
        language="fr",
        currency="EUR",
        incoterm="EXW",
        label="Offre",
        valid_from=None,
        valid_to=None,
        simulation=SimpleNamespace(market_params={}),  # no copper/fx
    )
    line = SimpleNamespace(final_price=Decimal("1"), product=SimpleNamespace(sku_code="P"))
    client = SimpleNamespace(name="C")
    data = build_tariff_xlsx(
        offer=offer,
        client=client,
        lines=[line],
        columns=["sku_code", "copper_base_price"],
        lang="fr",
    )
    ws = load_workbook(io.BytesIO(data)).active
    header_row = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Réf. SKU"
    )
    assert ws.cell(row=header_row + 1, column=2).value is None  # blank, no crash
