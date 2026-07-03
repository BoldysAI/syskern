"""Offer target-language resolution + pre-generation coverage (CDC §10.5.1)."""

from __future__ import annotations

from decimal import Decimal

import pytest
from rest_framework.test import APIClient

from apps.clients.models import Client
from apps.offers.models import Offer
from apps.offers.services.offer_i18n import (
    products_missing_language,
    resolve_product_description,
    resolve_product_designation,
)
from apps.offers.tasks import generate_tariff_offers_task, offer_export_path
from apps.products.models import Product
from apps.simulations.models import Simulation, SimulationLine

pytestmark = pytest.mark.django_db


@pytest.fixture()
def export_dir(tmp_path, monkeypatch):
    monkeypatch.setattr("apps.offers.tasks.EXPORT_DIR", tmp_path)
    return tmp_path


def _finalized_tariff(*, marketing: dict | None = None) -> Simulation:
    sim = Simulation.objects.create(
        label="Tarif",
        simulation_type="tariff",
        status="draft",
        market_params={"fx_eur_usd": "1.15", "fx_eur_rmb": "7.95"},
    )
    product = Product.objects.create(
        sku_code="CABLE-1",
        name="Câble",
        description_marketing=marketing if marketing is not None else {"fr": "Câble blindé"},
    )
    SimulationLine.objects.create(
        simulation=sim, product=product, pv_eur=Decimal("100.00"), status="ok"
    )
    Simulation.objects.filter(pk=sim.pk).update(status="finalized")
    sim.refresh_from_db()
    return sim


# ── resolve_product_description ──────────────────────────────────────────────


def test_resolve_uses_target_when_present():
    product = Product(description_marketing={"fr": "FR", "en": "EN"})
    text, fallback = resolve_product_description(product, "en", "description_marketing")
    assert (text, fallback) == ("EN", False)


def test_resolve_falls_back_to_fr():
    product = Product(description_marketing={"fr": "FR"})
    text, fallback = resolve_product_description(product, "en", "description_marketing")
    assert (text, fallback) == ("FR", True)


def test_resolve_empty_when_nothing():
    product = Product(description_marketing={})
    assert resolve_product_description(product, "en", "description_marketing") == ("", False)


def test_resolve_designation_uses_target_marketing():
    product = Product(
        sku_code="X",
        name="Câble",
        description_marketing={"fr": "Câble FR", "es": "Cable ES"},
    )
    text, fallback = resolve_product_designation(product, "es")
    assert (text, fallback) == ("Cable ES", False)


def test_resolve_designation_falls_back_to_fr_marketing():
    product = Product(
        sku_code="X",
        name="Câble",
        description_marketing={"fr": "Câble FR"},
    )
    text, fallback = resolve_product_designation(product, "es")
    assert (text, fallback) == ("Câble FR", True)


def test_products_missing_language():
    p_fr = Product(sku_code="A", description_marketing={"fr": "x"})
    p_en = Product(sku_code="B", description_marketing={"fr": "x", "en": "y"})
    missing = products_missing_language([p_fr, p_en], "en")
    assert [p.sku_code for p in missing] == ["A"]


# ── offer-coverage-check endpoint ────────────────────────────────────────────


def test_coverage_check_reports_fr_only_product():
    sim = _finalized_tariff(marketing={"fr": "Câble blindé"})
    api = APIClient()
    resp = api.post(
        f"/api/simulations/{sim.id}/offer-coverage-check/",
        {"language": "en"},
        format="json",
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["languages"] == ["en"]
    assert len(data["products"]) == 1
    assert data["products"][0]["sku_code"] == "CABLE-1"
    assert data["products"][0]["missing_langs"] == ["en"]


def test_coverage_check_clean_when_translated():
    sim = _finalized_tariff(marketing={"fr": "Câble", "en": "Cable"})
    api = APIClient()
    resp = api.post(
        f"/api/simulations/{sim.id}/offer-coverage-check/",
        {"language": "en"},
        format="json",
    )
    assert resp.status_code == 200
    assert resp.json()["products"] == []


def test_coverage_check_resolves_per_client_language():
    sim = _finalized_tariff(marketing={"fr": "Câble"})
    client_es = Client.objects.create(name="Cliente ES", preferred_language="es")
    api = APIClient()
    resp = api.post(
        f"/api/simulations/{sim.id}/offer-coverage-check/",
        {"client_ids": [str(client_es.id)], "language_per_client": True},
        format="json",
    )
    assert resp.status_code == 200
    assert resp.json()["languages"] == ["es"]
    assert len(resp.json()["products"]) == 1


# ── Per-client tariff language ───────────────────────────────────────────────


def test_tariff_uses_client_preferred_language(export_dir):
    sim = _finalized_tariff()
    client_en = Client.objects.create(name="Client EN", preferred_language="en")
    params = {
        "client_ids": [str(client_en.id)],
        "columns": [],
        "target_currency": "EUR",
        "language": "fr",
        "language_per_client": True,
        "expiration_date": None,
        "incoterm": "EXW",
        "label": "Tarif",
    }
    generate_tariff_offers_task.apply(args=[str(sim.id), params]).get()
    offer = Offer.objects.get(client_ids=[str(client_en.id)])
    assert offer.language == "en"
    assert offer_export_path(offer.id).is_file()


def test_tariff_falls_back_to_wizard_language(export_dir):
    sim = _finalized_tariff()
    client = Client.objects.create(name="Client", preferred_language="fr")
    params = {
        "client_ids": [str(client.id)],
        "columns": [],
        "target_currency": "EUR",
        "language": "es",
        "language_per_client": False,
        "expiration_date": None,
        "incoterm": "EXW",
        "label": "Tarif",
    }
    generate_tariff_offers_task.apply(args=[str(sim.id), params]).get()
    offer = Offer.objects.get(client_ids=[str(client.id)])
    assert offer.language == "es"


def test_tariff_excel_uses_spanish_designation(export_dir):
    sim = _finalized_tariff(
        marketing={"fr": "Câble blindé", "es": "Cable blindado"},
    )
    client = Client.objects.create(name="Cliente ES")
    params = {
        "client_ids": [str(client.id)],
        "columns": ["name", "sku_code"],
        "target_currency": "EUR",
        "language": "es",
        "language_per_client": False,
        "expiration_date": None,
        "incoterm": "EXW",
        "label": "Tarif ES",
    }
    generate_tariff_offers_task.apply(args=[str(sim.id), params]).get()
    offer = Offer.objects.get(client_ids=[str(client.id)])
    from openpyxl import load_workbook

    ws = load_workbook(offer_export_path(offer.id))["Tarifas"]
    header_row = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Designación"
    )
    assert ws.cell(row=header_row + 1, column=1).value == "Cable blindado"
