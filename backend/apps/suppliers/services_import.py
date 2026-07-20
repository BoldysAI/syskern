"""Flexible PO import services for the Fournisseurs module (Épic FEEDBACK 1).

The legacy import guessed the SKU / fournisseur / PO columns from header
aliases and broke whenever the Excel structure differed. This module replaces
that guessing with an **explicit, user-provided column mapping** and a two-phase
flow: a dry-run resolution (preview / synthesis) then an apply step.

``column_map`` maps a logical field to a **0-based column index** (not a header
label): ``{"sku": 0, "po": 3, "supplier": 1, ...}``. Indices are used rather than
labels so that columns **without a header name** are still mappable, and files
whose header row is not the first row work too (the header row is configurable
via ``header_row``, 1-based). Only ``sku`` and ``po`` are required. Money stays
``Decimal`` (AGENTS §5.1). Matching is by existing SKU / existing supplier only —
never creates a product, never creates a supplier.
"""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl

from apps.products.models import Currency, Incoterm, PriceChangeSource, Product, ProductSupplier

from .models import Supplier
from .services import record_po_change

if TYPE_CHECKING:
    from collections.abc import Iterable

# Row resolution statuses (also surfaced to the frontend synthesis step).
WILL_UPDATE = "will_update"
WILL_CREATE_LINK = "will_create_link"
UNCHANGED = "unchanged"
SKU_NOT_FOUND = "sku_not_found"
SUPPLIER_NOT_FOUND = "supplier_not_found"
INVALID_PO = "invalid_po"
NO_SUPPLIER = "no_supplier"
MISSING_SKU = "missing_sku"

# Statuses that mean "nothing will be written" (rejected / shown only).
_REJECTED_STATUSES = {
    SKU_NOT_FOUND,
    SUPPLIER_NOT_FOUND,
    INVALID_PO,
    NO_SUPPLIER,
    MISSING_SKU,
}

_STATUS_REASONS = {
    SKU_NOT_FOUND: "SKU introuvable en base",
    SUPPLIER_NOT_FOUND: "Fournisseur introuvable en base",
    INVALID_PO: "PO manquant ou invalide",
    NO_SUPPLIER: "Aucun fournisseur (ni colonne ni sélection)",
    MISSING_SKU: "SKU manquant",
}

# Logical fields the wizard can map. ``required`` gates the mapping step.
MAPPABLE_FIELDS = ("sku", "po", "supplier", "po_currency", "factory_code", "incoterm")
REQUIRED_FIELDS = ("sku", "po")

_PO_QUANTUM = Decimal("0.0001")


def _pad(rows: list[list], width: int) -> list[list]:
    return [list(r) + [""] * (width - len(r)) for r in rows]


def read_excel_headers(
    path: str | Path, *, header_row: int = 1, sample: int = 20
) -> tuple[list[str], list[list], int]:
    """Return ``(headers, sample_rows, column_count)`` with a **bounded** read.

    ``header_row`` is 1-based: rows before it are ignored (title/blank rows),
    the header labels are read from it, and ``sample`` data rows after it are
    returned. ``column_count`` is the widest row seen so **every** column —
    including unnamed ones — is addressable by index. openpyxl ``read_only`` is
    lazy: we stop after the sample (not a full parse — AGENTS §4).
    """
    header_idx = max(header_row, 1) - 1
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"Fichier Excel illisible : {exc}") from exc

    ws = wb.active
    headers: list[str] = []
    sample_rows: list[list] = []
    found = False
    for i, raw_row in enumerate(ws.iter_rows(values_only=True)):
        if i < header_idx:
            continue
        if i == header_idx:
            headers = [str(cell).strip() if cell is not None else "" for cell in raw_row]
            found = True
            continue
        sample_rows.append(["" if cell is None else str(cell) for cell in raw_row])
        if len(sample_rows) >= sample:
            break
    wb.close()

    if not found:
        raise ValueError(
            "Ligne d'en-tête introuvable (fichier vide ou en-tête au-delà du fichier)."
        )

    column_count = max([len(headers), *(len(r) for r in sample_rows)])
    headers = list(headers) + [""] * (column_count - len(headers))
    return headers, _pad(sample_rows, column_count), column_count


def _resolve_indices(column_map: dict, column_count: int) -> dict[str, int]:
    """Resolve each mapped logical field to a valid 0-based column index."""
    resolved: dict[str, int] = {}
    for field, raw in column_map.items():
        if field not in MAPPABLE_FIELDS or raw is None or raw == "":
            continue
        try:
            idx = int(raw)
        except (TypeError, ValueError):
            continue
        if 0 <= idx < column_count:
            resolved[field] = idx
    return resolved


def validate_column_map(column_map: dict, column_count: int) -> list[str]:
    """Return a list of FR error messages for an invalid mapping (empty = ok)."""
    errors: list[str] = []
    indices = _resolve_indices(column_map, column_count)
    for field in REQUIRED_FIELDS:
        if column_map.get(field) in (None, ""):
            errors.append(f"Le champ « {field.upper()} » doit être mappé.")
        elif field not in indices:
            errors.append(f"Colonne invalide dans le fichier pour « {field.upper()} ».")
    return errors


def _parse_price(raw: object) -> Decimal | None:
    if raw is None or str(raw).strip() == "":
        return None
    text = str(raw).strip().replace("\u00a0", "").replace(" ", "").replace(",", ".")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _clean_currency(raw: object) -> str | None:
    value = str(raw or "").strip().upper()
    return value if value in Currency.values else None


def _clean_incoterm(raw: object) -> str | None:
    value = str(raw or "").strip().upper()
    return value if value in Incoterm.values else None


def _cell(row: list, idx: int | None) -> object:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _iter_data_rows(path: str | Path, *, header_row: int = 1) -> tuple[list[list], int]:
    """Return ``(data_rows, column_count)`` — every row after ``header_row``.

    ``header_row`` is 1-based; rows before it (title/blank) and the header line
    itself are skipped. ``column_count`` is the widest row (header included) so
    unnamed / trailing columns stay addressable by index.
    """
    header_idx = max(header_row, 1) - 1
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"Fichier Excel illisible : {exc}") from exc
    ws = wb.active
    header_len = 0
    data_rows: list[list] = []
    found = False
    for i, raw_row in enumerate(ws.iter_rows(values_only=True)):
        if i < header_idx:
            continue
        if i == header_idx:
            header_len = len(raw_row)
            found = True
            continue
        data_rows.append(list(raw_row))
    wb.close()
    if not found:
        raise ValueError(
            "Ligne d'en-tête introuvable (fichier vide ou en-tête au-delà du fichier)."
        )
    column_count = max([header_len, *(len(r) for r in data_rows)])
    return data_rows, column_count


def _resolve_supplier_cache() -> dict[str, Supplier]:
    return {s.name.strip().lower(): s for s in Supplier.objects.all()}


def _build_line(
    *,
    row_number: int,
    sku: str,
    supplier_label: str,
    po_raw: object,
    status: str,
    old_price: Decimal | None = None,
    new_price: Decimal | None = None,
    po_currency: str | None = None,
    factory_code: str | None = None,
    incoterm: str | None = None,
) -> dict:
    return {
        "row": row_number,
        "sku": sku,
        "supplier": supplier_label,
        "po": "" if po_raw is None else str(po_raw),
        "status": status,
        "reason": _STATUS_REASONS.get(status, ""),
        "old_po_base_price": str(old_price) if old_price is not None else None,
        "new_po_base_price": str(new_price) if new_price is not None else None,
        "po_currency": po_currency,
        "factory_code": factory_code,
        "incoterm": incoterm,
    }


def _resolve_rows(
    path: str | Path,
    *,
    column_map: dict,
    default_supplier: Supplier | None,
    header_row: int = 1,
) -> tuple[list[dict], dict[str, int]]:
    """Resolve every data row against the DB. Pure read — no writes.

    Returns ``(lines, summary)``. Supplier resolution: mapped column wins per
    row; falls back to ``default_supplier``; if neither → ``no_supplier``.
    Matching is by existing SKU / existing supplier only.
    """
    data_rows, column_count = _iter_data_rows(path, header_row=header_row)
    indices = _resolve_indices(column_map, column_count)
    sku_idx = indices.get("sku")
    po_idx = indices.get("po")
    supplier_idx = indices.get("supplier")
    currency_idx = indices.get("po_currency")
    factory_idx = indices.get("factory_code")
    incoterm_idx = indices.get("incoterm")

    supplier_cache = _resolve_supplier_cache()
    lines: list[dict] = []
    counts: dict[str, int] = {}

    def bump(status: str) -> None:
        counts[status] = counts.get(status, 0) + 1

    for offset, row in enumerate(data_rows):
        row_number = header_row + 1 + offset  # 1-based Excel row of this data line
        sku = str(_cell(row, sku_idx) or "").strip()
        supplier_label = (
            str(_cell(row, supplier_idx) or "").strip() if supplier_idx is not None else ""
        )
        po_raw = _cell(row, po_idx)

        # Skip fully-empty trailing rows silently.
        if not sku and not supplier_label and (po_raw is None or str(po_raw).strip() == ""):
            continue

        if not sku:
            lines.append(
                _build_line(
                    row_number=row_number,
                    sku=sku,
                    supplier_label=supplier_label,
                    po_raw=po_raw,
                    status=MISSING_SKU,
                )
            )
            bump(MISSING_SKU)
            continue

        # Resolve supplier: mapped column wins, else default.
        supplier = None
        if supplier_label:
            supplier = supplier_cache.get(supplier_label.lower())
            if supplier is None:
                lines.append(
                    _build_line(
                        row_number=row_number,
                        sku=sku,
                        supplier_label=supplier_label,
                        po_raw=po_raw,
                        status=SUPPLIER_NOT_FOUND,
                    )
                )
                bump(SUPPLIER_NOT_FOUND)
                continue
        else:
            supplier = default_supplier

        if supplier is None:
            lines.append(
                _build_line(
                    row_number=row_number,
                    sku=sku,
                    supplier_label=supplier_label,
                    po_raw=po_raw,
                    status=NO_SUPPLIER,
                )
            )
            bump(NO_SUPPLIER)
            continue

        price = _parse_price(po_raw)
        if price is None:
            lines.append(
                _build_line(
                    row_number=row_number,
                    sku=sku,
                    supplier_label=supplier.name,
                    po_raw=po_raw,
                    status=INVALID_PO,
                )
            )
            bump(INVALID_PO)
            continue
        price = price.quantize(_PO_QUANTUM)

        product = Product.objects.filter(sku_code=sku).first()
        if product is None:
            lines.append(
                _build_line(
                    row_number=row_number,
                    sku=sku,
                    supplier_label=supplier.name,
                    po_raw=po_raw,
                    status=SKU_NOT_FOUND,
                )
            )
            bump(SKU_NOT_FOUND)
            continue

        factory_code = str(_cell(row, factory_idx) or "").strip() if factory_idx is not None else ""
        incoterm = _clean_incoterm(_cell(row, incoterm_idx)) if incoterm_idx is not None else None

        link = ProductSupplier.objects.filter(product=product, supplier=supplier).first()
        if link is None:
            currency = _clean_currency(_cell(row, currency_idx)) or supplier.currency_default
            lines.append(
                _build_line(
                    row_number=row_number,
                    sku=sku,
                    supplier_label=supplier.name,
                    po_raw=po_raw,
                    status=WILL_CREATE_LINK,
                    old_price=None,
                    new_price=price,
                    po_currency=currency,
                    factory_code=factory_code or None,
                    incoterm=incoterm,
                )
            )
            bump(WILL_CREATE_LINK)
            continue

        old_price = link.po_base_price
        currency = _clean_currency(_cell(row, currency_idx)) or link.po_currency
        status = UNCHANGED if old_price == price else WILL_UPDATE
        lines.append(
            _build_line(
                row_number=row_number,
                sku=sku,
                supplier_label=supplier.name,
                po_raw=po_raw,
                status=status,
                old_price=old_price,
                new_price=price,
                po_currency=currency,
                factory_code=factory_code or None,
                incoterm=incoterm,
            )
        )
        bump(status)

    return lines, counts


def _summary(counts: dict[str, int]) -> dict[str, int]:
    total = sum(counts.values())
    rejected = sum(counts.get(s, 0) for s in _REJECTED_STATUSES)
    return {
        "total": total,
        WILL_UPDATE: counts.get(WILL_UPDATE, 0),
        WILL_CREATE_LINK: counts.get(WILL_CREATE_LINK, 0),
        UNCHANGED: counts.get(UNCHANGED, 0),
        SKU_NOT_FOUND: counts.get(SKU_NOT_FOUND, 0),
        SUPPLIER_NOT_FOUND: counts.get(SUPPLIER_NOT_FOUND, 0),
        INVALID_PO: counts.get(INVALID_PO, 0),
        NO_SUPPLIER: counts.get(NO_SUPPLIER, 0),
        MISSING_SKU: counts.get(MISSING_SKU, 0),
        "rejected": rejected,
    }


def preview_import(
    path: str | Path,
    *,
    column_map: dict,
    default_supplier: Supplier | None,
    header_row: int = 1,
    row_cap: int = 1000,
    progress=None,
) -> dict:
    """Dry-run the import — returns ``{summary, lines}`` without any DB write.

    ``lines`` is capped to ``row_cap`` for the payload; ``summary`` counts all
    rows. ``progress`` (optional) is a callable ``(current, total)``.
    """
    lines, counts = _resolve_rows(
        path, column_map=column_map, default_supplier=default_supplier, header_row=header_row
    )
    if progress is not None:
        progress(len(lines), len(lines))
    return {"summary": _summary(counts), "lines": lines[:row_cap]}


def apply_import(
    path: str | Path,
    *,
    column_map: dict,
    default_supplier: Supplier | None,
    header_row: int = 1,
    progress=None,
) -> dict:
    """Apply the import: update / create links and write PO history.

    Re-resolves the rows then persists only the actionable ones. Each applied PO
    writes a ``SupplierPriceHistory`` entry (source=import). Returns
    ``{updated, created, rejected, total, rejected_rows}``.
    """
    lines, counts = _resolve_rows(
        path, column_map=column_map, default_supplier=default_supplier, header_row=header_row
    )
    supplier_cache = _resolve_supplier_cache()

    updated = 0
    created = 0
    actionable = [line for line in lines if line["status"] in {WILL_UPDATE, WILL_CREATE_LINK}]
    total_actionable = len(actionable)

    for done, line in enumerate(actionable, start=1):
        product = Product.objects.filter(sku_code=line["sku"]).first()
        supplier = supplier_cache.get(line["supplier"].strip().lower())
        if product is None or supplier is None:
            continue
        price = Decimal(line["new_po_base_price"]).quantize(_PO_QUANTUM)

        # Optional mapped fields (only override when provided in the file).
        mapped_currency = line.get("po_currency")
        mapped_factory = line.get("factory_code")
        mapped_incoterm = line.get("incoterm")

        link = ProductSupplier.objects.filter(product=product, supplier=supplier).first()
        if link is None:
            link = ProductSupplier(
                product=product,
                supplier=supplier,
                supplier_name=supplier.name,
                factory_code=mapped_factory or supplier.factory_code_default,
                po_currency=mapped_currency or supplier.currency_default,
                incoterm=mapped_incoterm or supplier.incoterm_default,
                incoterm_location=supplier.location,
                is_active=False,
            )
            was_created = True
        else:
            was_created = False
            if mapped_currency:
                link.po_currency = mapped_currency
            if mapped_factory:
                link.factory_code = mapped_factory
            if mapped_incoterm:
                link.incoterm = mapped_incoterm

        old_price = link.po_base_price
        link.po_base_price = price
        link.save()
        record_po_change(
            link, old_price=old_price, new_price=price, source=PriceChangeSource.IMPORT
        )
        if was_created:
            created += 1
        else:
            updated += 1
        if progress is not None:
            progress(done, total_actionable)

    rejected_lines = [line for line in lines if line["status"] in _REJECTED_STATUSES]
    return {
        "total": _summary(counts)["total"],
        "updated": updated,
        "created": created,
        "rejected": len(rejected_lines),
        "rejected_rows": rejected_lines[:1000],
    }


def rejected_report_rows(lines: Iterable[dict]) -> list[dict]:
    """Filter resolution lines down to the rejected ones for the Excel report."""
    return [line for line in lines if line["status"] in _REJECTED_STATUSES]
