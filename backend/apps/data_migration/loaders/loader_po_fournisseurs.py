"""Loader for the Symea Shanghai PO supplier file (CDC §8.4 — Étape 2).

Source file : UKN_RANGE_PRICES_MARCH 2026 - Q1 2026_v1 - OP Working file 97000.xlsx
Target sheet: 'PO & SC March 2026'
Header row  : 13 (0-based index 12)

What this loader does
---------------------
For each of the ~1 095 data rows it:
  1. Matches the product via sku_code (Unikkern Items code → rule 1) with
     factory_code + category as fallback (rules 2 / 3 of the §8.6 cascade).
  2. Enriches the Product record (descriptions, GTIN, HS code, copper weight,
     pallet qty, brand, active flag).
  3. Creates or updates a ProductSupplier record for the (product, factory_code)
     pair with the FOB USD price, copper-indexation flag, and incoterm.

Multi-supplier SKUs (87 out of 1 008 unique SKUs appear twice — one row per
manufacturing origin, e.g. China HT + Turkey KK).  Each row creates a
SEPARATE ProductSupplier, so dedup_key() returns None (no deduplication).

Metadata extraction (pre_run hook)
-----------------------------------
The file header (rows 1–12) carries:
  - 3mm copper base price  → ProductSupplier.copper_base_price
  - Supplier code list     → factory_code ↔ supplier_name mapping
These are read once before the main loop via pre_run().

Conservative update policy
---------------------------
Product fields are written only when the source value is non-null and not an
Excel error string (so existing good data from Odoo is never overwritten with
#N/A / #ERROR!).

ProductSupplier is upserted via get_or_create(product=…, factory_code=…).
Rows without an Internal Code suffix use a synthetic factory_code
``?`` + Supplier (max 15 chars) so distinct suppliers never share the empty
``factory_code`` key.  Rows with neither suffix nor Supplier column are
quarantined (InvalidRowError).

is_active is left at its current DB value — Olivier activates the preferred
supplier manually after migration (or the derivation step does it).
"""
from __future__ import annotations

import logging
from decimal import Decimal, InvalidOperation

import openpyxl
import pandas as pd

from apps.core.models import Currency
from apps.products.models import Incoterm, MigrationSource, Product, ProductSupplier

from .base import BaseExcelLoader
from .exceptions import InvalidRowError
from .io import coerce_decimal, coerce_int, coerce_str, row_to_raw
from .types import LoaderConfig, MatchHint, NormalizedRow, RowOutcome

logger = logging.getLogger(__name__)

# Sheet containing supplier ID → numeric factory code mapping
_SUPPLIER_SHEET = "Supplier code list"
# Default incoterm for this file's FOB pricing
_INCOTERM = Incoterm.FOB
# Currency for FOB prices in this file
_PO_CURRENCY = Currency.USD
# Name of the trading company that sourced this price list
_TRADING_COMPANY = "Symea Shanghai"

# Excel error strings that should be treated as missing values
_EXCEL_ERRORS = {"#REF!", "#N/A", "#DIV/0!", "#VALUE!", "#NAME?", "#NULL!", "#NUM!", "#ERROR!"}


def _clean(value: object) -> str | None:
    """Return a non-empty, non-error string or None."""
    s = coerce_str(value)
    if s is None or s in _EXCEL_ERRORS:
        return None
    return s


class POFournisseursLoader(BaseExcelLoader):
    """Loader for the Symea Shanghai PO supplier Excel file."""

    migration_source = MigrationSource.EXCEL_PRICING

    def __init__(self) -> None:
        # Populated by pre_run()
        self._copper_base_price: Decimal | None = None
        # factory_code (str) → supplier short name (str), e.g. {"17": "ZD", "21": "HT"}
        self._factory_to_name: dict[str, str] = {}

    # ── pre_run hook ──────────────────────────────────────────────────────────

    def pre_run(self, config: LoaderConfig) -> None:
        """Read per-file metadata: copper base price and supplier code list."""
        wb = openpyxl.load_workbook(config.file_path, data_only=True, read_only=True)
        self._read_copper_base(wb)
        self._read_supplier_codes(wb)
        logger.info(
            "Metadata: copper_base=%s, suppliers=%s",
            self._copper_base_price,
            self._factory_to_name,
        )

    def _read_copper_base(self, wb: openpyxl.Workbook) -> None:
        """Parse the '3mm copper base' value from the PO sheet header rows."""
        ws = wb.worksheets[0]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 12:
                break
            vals = list(row)
            for j, cell in enumerate(vals):
                if str(cell).strip() == "3mm copper base" and j + 1 < len(vals):
                    raw = vals[j + 1]
                    try:
                        self._copper_base_price = Decimal(str(raw))
                    except (InvalidOperation, TypeError):
                        logger.warning("Could not parse 3mm copper base value: %r", raw)
                    return

    def _read_supplier_codes(self, wb: openpyxl.Workbook) -> None:
        """Build factory_code → supplier_name mapping from 'Supplier code list' sheet."""
        if _SUPPLIER_SHEET not in wb.sheetnames:
            logger.warning("Sheet %r not found; supplier names will be unknown.", _SUPPLIER_SHEET)
            return
        ws = wb[_SUPPLIER_SHEET]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:  # header row
                continue
            if not row or row[1] is None:
                continue
            supplier_id = str(row[1]).strip()  # e.g. "ZD", "HT"
            raw_code = row[2]               # e.g. 17.0, "E02"
            if raw_code is None:
                continue
            # Normalise numeric codes to strings without trailing .0
            try:
                factory_code = str(int(float(str(raw_code))))
            except (ValueError, TypeError):
                factory_code = str(raw_code).strip()
            if factory_code:
                self._factory_to_name[factory_code] = supplier_id

    # ── Column mapping ────────────────────────────────────────────────────────

    def column_mapping(self) -> dict[str, str]:
        return {
            "Unikkern Items code": "sku_code",
            "Internal Code": "internal_code",
            "Brand": "brand",
            "Active France Y/N": "active_france",
            "Active Export Y/N": "active_export",
            "Univers": "universe",
            "Range": "range",
            "Sub-Range": "sub_range",
            "Type": "cable_type",
            "AWG/SIZE": "awg",
            "item code": "item_code",
            "Description En": "description_en",
            "Description fr": "description_fr",
            "Tag": "cpr_tag",
            "Origin": "origin",
            "Cu (kg/km)": "copper_weight",
            "QTY/Pallet": "pallet_qty",
            "Global Trade Item Number (GTIN)": "gtin",
            "HS Code": "hs_code",
            "MOQ ": "moq",
            "Supplier Payment term": "payment_term",
            "Lead time ": "lead_time",
            "Distributor SC FOB Price (Usd/km)": "fob_price_usd",
            "MB% FOB Symea": "symea_margin_rate",
            "Supplier": "supplier_code",
        }

    def required_columns(self) -> set[str]:
        return {"sku_code", "fob_price_usd"}

    # ── Row normalisation ─────────────────────────────────────────────────────

    def normalize_row(self, raw: pd.Series) -> NormalizedRow:
        sku = _clean(raw.get("sku_code"))
        internal_code = _clean(raw.get("internal_code"))

        # Derive factory_code from Internal Code suffix (e.g. "KCFF6A4-21" → "21")
        factory_code: str | None = None
        if internal_code and "-" in internal_code:
            factory_code = internal_code.rsplit("-", 1)[-1].strip() or None

        fob_str = coerce_decimal(raw.get("fob_price_usd"))

        # Build description_marketing merge dict preserving existing keys
        desc_en = _clean(raw.get("description_en"))
        desc_fr = _clean(raw.get("description_fr"))

        copper_str = coerce_decimal(raw.get("copper_weight"))
        copper = Decimal(copper_str) if copper_str else None

        pallet = coerce_int(raw.get("pallet_qty"))

        active_france = str(raw.get("active_france") or "").strip().lower() == "yes"
        active_export = str(raw.get("active_export") or "").strip().lower() == "yes"

        # Symea margin rate: numeric 0–1 (e.g. 0.09)
        margin_str = coerce_decimal(raw.get("symea_margin_rate"))

        data = {
            "sku_code": sku,
            "internal_code": internal_code,
            "factory_code": factory_code,
            "brand": _clean(raw.get("brand")),
            "is_active": active_france or active_export,
            "universe": _clean(raw.get("universe")),
            "range": _clean(raw.get("range")),
            "sub_range": _clean(raw.get("sub_range")),
            "item_code": _clean(raw.get("item_code")),
            "description_en": desc_en,
            "description_fr": desc_fr,
            "gtin": _clean(raw.get("gtin")),
            "hs_code": _clean(raw.get("hs_code")),
            "copper_weight": copper,
            "pallet_qty": pallet,
            "fob_price_usd": Decimal(fob_str) if fob_str else None,
            "supplier_code": _clean(raw.get("supplier_code")),
            "origin": _clean(raw.get("origin")),
            "payment_term": _clean(raw.get("payment_term")),
            "moq": _clean(raw.get("moq")),
            "symea_margin_rate": Decimal(margin_str) if margin_str else None,
        }

        return NormalizedRow(data=data, raw=row_to_raw(raw))

    # ── Match hint ────────────────────────────────────────────────────────────

    def build_match_hint(self, row: NormalizedRow) -> MatchHint:
        sku = row.data.get("sku_code")

        # parent_reference for rule 2: sku_code itself is the parent
        # factory_code for rules 2 & 3: derived from Internal Code
        factory = row.data.get("factory_code")

        # category hint: "universe|range|sub_range|" for rule 3
        universe = row.data.get("universe") or ""
        rng = row.data.get("range") or ""
        sub = row.data.get("sub_range") or ""
        category = f"{universe.upper()}|{rng.upper()}|{sub.upper()}|" if any([universe, rng, sub]) else None

        return MatchHint(
            sku_code=sku,
            parent_reference=sku,  # for this file, sku_code == parent_reference
            factory_code=factory,
            category=category,
        )

    # ── DB update ─────────────────────────────────────────────────────────────

    def apply_update(self, product: Product, row: NormalizedRow) -> RowOutcome:
        """Enrich the Product and upsert the ProductSupplier."""
        self._update_product(product, row)
        self._upsert_supplier(product, row)
        return RowOutcome(row_number=0, matched=True, updated=True, quarantined=False)

    def _update_product(self, product: Product, row: NormalizedRow) -> None:
        d = row.data
        changed = False

        def _set(field: str, value: object) -> None:
            nonlocal changed
            if value is not None and getattr(product, field) != value:
                setattr(product, field, value)
                changed = True

        _set("brand", d.get("brand"))
        _set("item_code", d.get("item_code"))
        _set("is_active", d.get("is_active"))

        # Hierarchy: only fill if currently empty
        for field in ("universe", "range", "sub_range"):
            if not getattr(product, field):
                _set(field, d.get(field))

        # Copper weight
        _set("copper_weight_kg_per_unit", d.get("copper_weight"))

        # is_copper_indexed derived from copper weight presence
        if d.get("copper_weight") is not None:
            new_indexed = d["copper_weight"] > 0
            if product.is_copper_indexed != new_indexed:
                product.is_copper_indexed = new_indexed
                changed = True

        _set("pallet_qty", d.get("pallet_qty"))

        # GTIN & HS code: only set if valid (non-null, non-error)
        _set("gtin", d.get("gtin"))
        _set("hs_code", d.get("hs_code"))

        # Multilingual descriptions: merge into existing JSONB dict (first-wins)
        desc = dict(product.description_marketing or {})
        updated_desc = False
        if d.get("description_en") and not desc.get("en"):
            desc["en"] = d["description_en"]
            updated_desc = True
        if d.get("description_fr") and not desc.get("fr"):
            desc["fr"] = d["description_fr"]
            updated_desc = True
        if updated_desc and desc != product.description_marketing:
            product.description_marketing = desc
            changed = True

        if changed:
            product.save()

    @staticmethod
    def _storage_factory_key(d: dict[str, object]) -> str:
        """Key for get_or_create(product, factory_code=…).

        Real suffixes from Internal Code are used as-is.  When the suffix is
        missing, ``?`` + trimmed Supplier column keeps rows disjoint; bare
        ``""`` would merge unrelated suppliers on the same product.
        """
        raw = d.get("factory_code")
        if isinstance(raw, str) and raw.strip():
            return raw.strip()
        sup = d.get("supplier_code")
        sup_s = sup.strip() if isinstance(sup, str) else ""
        if not sup_s:
            raise InvalidRowError(
                "factory_code is missing (Internal Code has no factory suffix) and "
                "Supplier is empty — cannot upsert supplier without a disambiguating key"
            )
        return f"?{sup_s[:15]}"

    def _upsert_supplier(self, product: Product, row: NormalizedRow) -> None:
        d = row.data
        factory_key = self._storage_factory_key(d)
        raw_fc = d.get("factory_code")
        parsed_factory = raw_fc.strip() if isinstance(raw_fc, str) and raw_fc.strip() else None
        supplier_code = d.get("supplier_code") or ""
        supplier_code = supplier_code.strip() if isinstance(supplier_code, str) else ""
        if not supplier_code and parsed_factory:
            supplier_code = self._factory_to_name.get(parsed_factory, "")

        fob_price = d.get("fob_price_usd")
        if fob_price is None:
            raise InvalidRowError("fob_price_usd is null — cannot create supplier record")

        copper_weight = d.get("copper_weight")
        is_copper = copper_weight is not None and copper_weight > 0

        supplier_notes_parts = []
        if parsed_factory is None:
            supplier_notes_parts.append(
                "factory_code: derived from Supplier column (Internal Code suffix missing)"
            )
        if d.get("payment_term"):
            supplier_notes_parts.append(f"Payment: {d['payment_term']}")
        if d.get("moq"):
            supplier_notes_parts.append(f"MOQ: {d['moq']}")
        if d.get("symea_margin_rate") is not None:
            supplier_notes_parts.append(f"Symea margin: {d['symea_margin_rate']}")
        notes = " | ".join(supplier_notes_parts)

        supplier, created = ProductSupplier.objects.get_or_create(
            product=product,
            factory_code=factory_key,
            defaults={
                "supplier_name": supplier_code or _TRADING_COMPANY,
                "is_active": False,
            },
        )

        # Always update pricing fields — this is the enrichment step
        supplier.supplier_name = supplier_code or _TRADING_COMPANY
        supplier.po_base_price = fob_price
        supplier.po_currency = _PO_CURRENCY
        supplier.incoterm = _INCOTERM
        supplier.incoterm_location = d.get("origin") or ""
        supplier.is_copper_indexed = is_copper
        if is_copper and self._copper_base_price is not None:
            supplier.copper_base_price = self._copper_base_price
        if notes:
            supplier.notes = notes
        supplier.save()

        action = "created" if created else "updated"
        logger.debug(
            "ProductSupplier %s: product=%s factory=%s supplier=%s fob=%s",
            action,
            product.sku_code,
            factory_key,
            supplier.supplier_name,
            fob_price,
        )
