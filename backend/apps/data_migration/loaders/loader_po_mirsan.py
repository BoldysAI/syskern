"""Loader for the Mirsan racks & network infrastructure PO file (CDC §8.4 — Étape 2).

Source file : MIRSAN_PRICES_LIST_2025_2026 (RACKS).xlsx
Supplier    : Mirsan (factory code: MIRSAN, origin: Turkey)
Products    : Network cabinets, open racks, wall cabinets, accessories

Multi-sheet strategy
--------------------
The file contains 4 processable sheets with slightly different column layouts.
Rather than running the loader once per sheet, ``run()`` is overridden to
iterate over all sheets in a single pass, sharing one ``ProductMatcher`` index
and accumulating into a single ``LoaderReport``.

Each sheet is normalised to the same canonical column schema before being
passed to ``_run_dataframe()``.

Sheet list and column profiles
-------------------------------
| Sheet                | header_row (0-based) | SKU col   | Price col            |
|----------------------|----------------------|-----------|----------------------|
| START CABINETS       | 5                    | col 1     | SYSKERN DDP (col 16) |
| GRID CABINETS        | 4                    | col 1     | SYSKERN DDP (col 13) |
| RACKS & OPEN RACKS   | 4                    | col 1     | SYSKERN DDP (col 13) |
| ACCESSORIES 19       | 3                    | col 2     | SYMEA DDP  (col  9)  |

Price policy
------------
We use the **SYSKERN DDP price** (delivered price Syskern pays) as the
``po_base_price`` with ``incoterm=DDP``.  For sheets where that column is
missing or has '-', we fall back to the "TG" (target) column.
The ``po_currency`` is EUR throughout.

No copper indexation — racks do not contain copper cable.
``is_active`` is left at its DB value (same conservative policy as the cable loaders).
"""
from __future__ import annotations

import logging
import time
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import ClassVar

import openpyxl
import pandas as pd

from apps.products.models import Currency, Incoterm, MigrationSource, Product, ProductSupplier

from .base import BaseExcelLoader
from .exceptions import InvalidRowError
from .io import coerce_decimal, coerce_str, row_to_raw
from .matching import ProductMatcher
from .types import LoaderConfig, LoaderReport, MatchHint, NormalizedRow, RowOutcome

logger = logging.getLogger(__name__)

_SUPPLIER_NAME = "Mirsan"
_FACTORY_CODE = "MIRSAN"
_INCOTERM = Incoterm.DDP
_PO_CURRENCY = Currency.EUR
_ORIGIN = "Turkey"


@dataclass(frozen=True)
class _SheetProfile:
    """Per-sheet column layout for the MIRSAN price list."""

    sheet_name: str
    header_row: int      # 0-based
    sku_col: int         # UKN product code column index
    item_code_col: int   # UKN number code column index (−1 = absent)
    desc_col: int        # Designation / Description column index
    mirsan_col: int      # Mirsan product code column index
    dim_col: int         # Dimensions column index (−1 = absent)
    price_col: int       # Best price column index (SYSKERN DDP or TG fallback)


# Profiles derived from manual inspection (see module docstring).
_SHEET_PROFILES: ClassVar[list[_SheetProfile]] = [
    _SheetProfile(
        sheet_name="START CABINETS",
        header_row=5,
        sku_col=1,
        item_code_col=2,
        desc_col=4,
        mirsan_col=0,
        dim_col=3,
        price_col=16,   # SYSKERN DDP Prices (latest)
    ),
    _SheetProfile(
        sheet_name="GRID CABINETS",
        header_row=4,
        sku_col=1,
        item_code_col=2,
        desc_col=4,
        mirsan_col=0,
        dim_col=3,
        price_col=13,   # SYSKERN DDP Prices
    ),
    _SheetProfile(
        sheet_name="RACKS & OPEN RACKS",
        header_row=4,
        sku_col=1,
        item_code_col=2,
        desc_col=0,
        mirsan_col=3,
        dim_col=-1,
        price_col=13,   # SYSKERN DDP prices
    ),
    _SheetProfile(
        sheet_name="ACCESSORIES 19",
        header_row=3,
        sku_col=2,
        item_code_col=3,
        desc_col=4,
        mirsan_col=0,
        dim_col=-1,
        price_col=9,    # SYMEA DDP Prices (best available; no SYSKERN DDP col)
    ),
]

# Canonical column names after normalisation
_CANON_SKU = "sku_code"
_CANON_ITEM = "item_code"
_CANON_DESC = "description"
_CANON_MIRSAN = "mirsan_code"
_CANON_DIM = "dimensions"
_CANON_PRICE = "price_eur"
_CANON_SHEET = "__sheet__"


def _is_valid_ukn_code(value: object) -> bool:
    """Return True only for actual UKN product codes (K + 2 uppercase letters…)."""
    s = coerce_str(value)
    return bool(s and len(s) >= 4 and s[0] == "K" and s[1:3].isupper())


def _read_sheet_as_df(wb: openpyxl.Workbook, profile: _SheetProfile) -> pd.DataFrame:
    """Read one sheet into a normalised DataFrame using the profile's column positions."""
    if profile.sheet_name not in wb.sheetnames:
        logger.warning("Sheet %r not found in workbook — skipping.", profile.sheet_name)
        return pd.DataFrame()

    ws = wb[profile.sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))

    records = []
    for row in all_rows[profile.header_row + 1 :]:
        sku = coerce_str(row[profile.sku_col] if profile.sku_col < len(row) else None)
        if not _is_valid_ukn_code(sku):
            continue  # skip header repeats, section titles, empty rows

        n = len(row)
        price_raw = row[profile.price_col] if profile.price_col < n else None
        # RACKS sheet stores TG as "$120,00" string — strip to numeric
        if isinstance(price_raw, str):
            price_raw = price_raw.replace("$", "").replace(",", ".").strip() or None

        item_raw = row[profile.item_code_col] if profile.item_code_col >= 0 < n else None
        desc_raw = row[profile.desc_col] if profile.desc_col < n else None
        mirsan_raw = row[profile.mirsan_col] if profile.mirsan_col < n else None
        dim_raw = row[profile.dim_col] if profile.dim_col >= 0 < n else None

        records.append(
            {
                _CANON_SKU: sku,
                _CANON_ITEM: coerce_str(item_raw),
                _CANON_DESC: coerce_str(desc_raw),
                _CANON_MIRSAN: coerce_str(mirsan_raw),
                _CANON_DIM: coerce_str(dim_raw),
                _CANON_PRICE: price_raw,
                _CANON_SHEET: profile.sheet_name,
            }
        )

    return pd.DataFrame(records)


class MirsanLoader(BaseExcelLoader):
    """Loader for the Mirsan racks & infrastructure PO Excel file."""

    migration_source = MigrationSource.EXCEL_PRICING

    # ── Column mapping (already canonical — no rename needed) ─────────────────

    def column_mapping(self) -> dict[str, str]:
        return {}  # sheets are pre-normalised in run()

    def required_columns(self) -> set[str]:
        return {_CANON_SKU, _CANON_PRICE}

    # ── Override run() to iterate over all sheets ─────────────────────────────

    def run(self, config: LoaderConfig) -> LoaderReport:  # type: ignore[override]
        start = time.monotonic()
        source_name = Path(config.file_path).name
        logger.info("Starting MirsanLoader on %s (dry_run=%s)", source_name, config.dry_run)

        wb = openpyxl.load_workbook(config.file_path, data_only=True, read_only=True)
        matcher = ProductMatcher()
        report = LoaderReport(source_file=source_name, sheet_name="(all sheets)", dry_run=config.dry_run)

        for profile in _SHEET_PROFILES:
            df = _read_sheet_as_df(wb, profile)
            if df.empty:
                logger.info("  Sheet %r: no UKN rows — skipped.", profile.sheet_name)
                continue
            logger.info("  Sheet %r: %d UKN rows to process.", profile.sheet_name, len(df))
            # Use a per-sheet config clone with batch_size and dry_run preserved
            sheet_config = LoaderConfig(
                file_path=config.file_path,
                sheet_name=profile.sheet_name,
                header_row=profile.header_row,
                batch_size=config.batch_size,
                dry_run=config.dry_run,
            )
            self._run_dataframe(df, sheet_config, matcher, report, source_name)

        report.duration_seconds = time.monotonic() - start
        logger.info("Finished MirsanLoader:\n%s", report)
        return report

    # ── Row normalisation ─────────────────────────────────────────────────────

    def normalize_row(self, raw: pd.Series) -> NormalizedRow:
        sku = coerce_str(raw.get(_CANON_SKU))
        price_str = coerce_decimal(raw.get(_CANON_PRICE))

        desc = coerce_str(raw.get(_CANON_DESC))
        dim = coerce_str(raw.get(_CANON_DIM))
        full_desc = f"{desc} — {dim}" if dim and desc else desc or dim

        data = {
            "sku_code": sku,
            "item_code": coerce_str(raw.get(_CANON_ITEM)),
            "description": full_desc,
            "mirsan_code": coerce_str(raw.get(_CANON_MIRSAN)),
            "price_eur": Decimal(price_str) if price_str else None,
            "sheet": coerce_str(raw.get(_CANON_SHEET)),
        }
        return NormalizedRow(data=data, raw=row_to_raw(raw))

    # ── Match hint ────────────────────────────────────────────────────────────

    def build_match_hint(self, row: NormalizedRow) -> MatchHint:
        sku = row.data.get("sku_code")
        return MatchHint(
            sku_code=sku,
            parent_reference=sku,
            factory_code=_FACTORY_CODE,
            category=None,
        )

    # ── DB update ─────────────────────────────────────────────────────────────

    def apply_update(self, product: Product, row: NormalizedRow) -> RowOutcome:
        self._update_product(product, row)
        self._upsert_supplier(product, row)
        return RowOutcome(row_number=0, matched=True, updated=True, quarantined=False)

    def _update_product(self, product: Product, row: NormalizedRow) -> None:
        d = row.data
        changed = False

        def _set(field: str, value: object) -> None:
            nonlocal changed
            if value is not None and getattr(product, field) != value:
                setattr(product, field, value)
                changed = True

        _set("item_code", d.get("item_code"))

        desc = d.get("description")
        if desc:
            existing = dict(product.description_marketing or {})
            if existing.get("en") != desc:
                existing["en"] = desc
                product.description_marketing = existing
                changed = True

        if changed:
            product.save()

    def _upsert_supplier(self, product: Product, row: NormalizedRow) -> None:
        d = row.data
        price = d.get("price_eur")
        if price is None or not isinstance(price, Decimal):
            raise InvalidRowError(f"price_eur missing for {d.get('sku_code')!r}")

        notes_parts = []
        if d.get("mirsan_code"):
            notes_parts.append(f"Mirsan ref: {d['mirsan_code']}")
        if d.get("sheet"):
            notes_parts.append(f"Sheet: {d['sheet']}")

        supplier, created = ProductSupplier.objects.get_or_create(
            product=product,
            factory_code=_FACTORY_CODE,
            defaults={
                "supplier_name": _SUPPLIER_NAME,
                "is_active": False,
            },
        )
        supplier.supplier_name = _SUPPLIER_NAME
        supplier.po_base_price = price
        supplier.po_currency = _PO_CURRENCY
        supplier.incoterm = _INCOTERM
        supplier.incoterm_location = _ORIGIN
        supplier.is_copper_indexed = False
        if notes_parts:
            supplier.notes = " | ".join(notes_parts)
        supplier.save()

        logger.debug(
            "ProductSupplier %s: product=%s price=%.2f EUR DDP",
            "created" if created else "updated",
            product.sku_code,
            price,
        )
