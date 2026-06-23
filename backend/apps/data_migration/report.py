"""Final post-migration report (CDC §8.8).

Cross-validation document produced for Olivier before go-live. Builds:

* a multi-tab Excel workbook (``migration_report_<date>.xlsx``):
  Synthèse, Fournisseurs, Attributs, Quarantaine, Dérivations, Simulation;
* a short plain-text email body with the headline figures.

The "créés vs mis à jour" counts and the run duration come from the
orchestrator's resume checkpoint (``MIGRATION_STATE_FILE``) when present — that
is the only place those per-step numbers are recorded; everything else is
aggregated live from the DB. The module is pure (no I/O side effects beyond
``generate_report`` writing the file) so :func:`build_report_data` is unit-testable.
"""

from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path

from django.conf import settings
from django.db.models import Avg, Count
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from apps.attributes.models import AttributeRegistry
from apps.clients.models import Client
from apps.products.models import Product, ProductSupplier
from apps.simulations.models import SimulationLine, SimulationStatus

from .models import MigrationUnmatched

logger = logging.getLogger("apps.data_migration.report")

_BLANK_LABEL = "(non renseigné)"
_HEADER_FILL = PatternFill("solid", fgColor="0F2137")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=13)


# ── Data aggregation ─────────────────────────────────────────────────────────


def _read_checkpoint(state_file: str | Path | None) -> dict | None:
    path = Path(state_file) if state_file else Path(settings.MIGRATION["STATE_FILE"])
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None


def _count_by(model, field: str) -> list[dict]:
    rows = model.objects.values(field).annotate(count=Count("id")).order_by(field)
    return [{"value": (r[field] or _BLANK_LABEL), "count": r["count"]} for r in rows]


def build_report_data(
    *, state_file: str | Path | None = None, when: datetime | None = None
) -> dict:
    """Aggregate every figure the report needs into a plain dict."""
    products_by_source = _count_by(Product, "migration_source")

    suppliers_by_source = [
        {"value": (r["product__migration_source"] or _BLANK_LABEL), "count": r["count"]}
        for r in (
            ProductSupplier.objects.values("product__migration_source")
            .annotate(count=Count("id"))
            .order_by("product__migration_source")
        )
    ]

    # Quarantine breakdown (CDC §8.7).
    q = MigrationUnmatched.objects.all()
    q_total = q.count()
    q_resolved = q.filter(resolved_at__isnull=False).count()
    by_reason = {
        r["reason"]: r["count"]
        for r in q.values("reason").annotate(count=Count("id")).order_by("reason")
    }
    by_source_file = {
        r["source_file"]: r["count"]
        for r in q.values("source_file").annotate(count=Count("id")).order_by("source_file")
    }
    quarantine_rows = list(
        q.order_by("source_file", "source_row_number").values(
            "source_file",
            "source_row_number",
            "reason",
            "resolved_at",
            "resolved_by",
            "resolution_notes",
        )
    )

    # Derivations summary (CDC §8.5).
    derivations = {
        "copper_indexed": Product.objects.filter(is_copper_indexed=True).count(),
        "base_unit": _count_by(Product, "base_unit"),
        "factory_code_filled": Product.objects.exclude(factory_code="").count(),
        "parent_reference_filled": Product.objects.exclude(parent_reference="").count(),
    }

    # Initial-simulation stats — PV moyen par gamme over finalized simulations.
    finalized_lines = SimulationLine.objects.filter(simulation__status=SimulationStatus.FINALIZED)
    pv_by_range = [
        {
            "range": (r["product__range"] or _BLANK_LABEL),
            "avg_pv_eur": float(r["avg_pv"]) if r["avg_pv"] is not None else None,
            "count": r["count"],
        }
        for r in (
            finalized_lines.values("product__range")
            .annotate(avg_pv=Avg("pv_eur"), count=Count("id"))
            .order_by("product__range")
        )
    ]
    simulations = {
        "available": bool(pv_by_range),
        "finalized_line_count": finalized_lines.count(),
        "pv_by_range": pv_by_range,
        "anomalies": finalized_lines.filter(status__in=["warning", "error"]).count(),
    }

    checkpoint = _read_checkpoint(state_file)

    return {
        "generated_at": (when or datetime.now()).isoformat(timespec="seconds"),
        "totals": {
            "products": Product.objects.count(),
            "suppliers": ProductSupplier.objects.count(),
            "clients": Client.objects.count(),
            "attributes": AttributeRegistry.objects.count(),
            "quarantine": q_total,
        },
        "products_by_source": products_by_source,
        "suppliers_by_source": suppliers_by_source,
        "attributes": {
            "total": AttributeRegistry.objects.count(),
            "by_category": _count_by(AttributeRegistry, "category"),
        },
        "quarantine": {
            "total": q_total,
            "resolved": q_resolved,
            "unresolved": q_total - q_resolved,
            "by_reason": by_reason,
            "by_source_file": by_source_file,
            "rows": quarantine_rows,
        },
        "derivations": derivations,
        "simulations": simulations,
        "run": checkpoint,  # None when no checkpoint on disk
    }


# ── Excel rendering ──────────────────────────────────────────────────────────


def _style_header(ws: Worksheet, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT


def _write_table(ws: Worksheet, start_row: int, headers: list[str], rows: list[list]) -> int:
    """Write a header + rows block starting at *start_row*. Returns next free row."""
    for col, head in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col, value=head)
    _style_header(ws, start_row, len(headers))
    r = start_row + 1
    for row_vals in rows:
        for col, val in enumerate(row_vals, start=1):
            ws.cell(row=r, column=col, value=val)
        r += 1
    return r + 1  # one blank spacer row


def build_workbook(data: dict) -> Workbook:
    """Render the report dict into a multi-tab openpyxl workbook."""
    wb = Workbook()

    # ── Synthèse ──
    ws = wb.active
    ws.title = "Synthèse"
    ws["A1"] = "Rapport de migration — Synthèse"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"Généré le {data['generated_at']}"
    row = 4
    row = _write_table(
        ws,
        row,
        ["Indicateur", "Valeur"],
        [
            ["Produits", data["totals"]["products"]],
            ["Fournisseurs", data["totals"]["suppliers"]],
            ["Clients", data["totals"]["clients"]],
            ["Attributs (registre)", data["totals"]["attributes"]],
            ["Lignes en quarantaine", data["totals"]["quarantine"]],
        ],
    )
    row = _write_table(
        ws,
        row,
        ["Produits par source", "Nombre"],
        [[s["value"], s["count"]] for s in data["products_by_source"]],
    )
    run = data.get("run")
    if run:
        created = sum(s.get("created", 0) for s in run.get("steps", []))
        updated = sum(s.get("updated", 0) for s in run.get("steps", []))
        duration = sum(s.get("duration_seconds", 0) for s in run.get("steps", []))
        _write_table(
            ws,
            row,
            ["Exécution (checkpoint)", "Valeur"],
            [
                ["Statut", run.get("status", "—")],
                ["Total créés", created],
                ["Total mis à jour", updated],
                ["Durée cumulée (s)", round(duration, 2)],
            ],
        )
    else:
        ws.cell(
            row=row,
            column=1,
            value="Aucun checkpoint d'exécution trouvé (créés/maj indisponibles).",
        )

    # ── Fournisseurs ──
    ws = wb.create_sheet("Fournisseurs")
    ws["A1"] = "Fournisseurs créés par source produit"
    ws["A1"].font = _TITLE_FONT
    _write_table(
        ws,
        3,
        ["Source produit", "Fournisseurs"],
        [[s["value"], s["count"]] for s in data["suppliers_by_source"]] or [["—", 0]],
    )

    # ── Attributs ──
    ws = wb.create_sheet("Attributs")
    ws["A1"] = "Attributs distincts du registre"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = f"Total : {data['attributes']['total']}"
    _write_table(
        ws,
        4,
        ["Catégorie", "Nombre"],
        [[c["value"], c["count"]] for c in data["attributes"]["by_category"]] or [["—", 0]],
    )

    # ── Quarantaine ──
    ws = wb.create_sheet("Quarantaine")
    ws["A1"] = "Lignes en quarantaine (CDC §8.7)"
    ws["A1"].font = _TITLE_FONT
    qd = data["quarantine"]
    ws["A2"] = f"Total {qd['total']} · résolues {qd['resolved']} · non résolues {qd['unresolved']}"
    row = _write_table(
        ws, 4, ["Raison", "Nombre"], [[k, v] for k, v in qd["by_reason"].items()] or [["—", 0]]
    )
    _write_table(
        ws,
        row,
        ["Fichier source", "Ligne", "Raison", "Résolue le", "Résolue par", "Notes"],
        [
            [
                r["source_file"],
                r["source_row_number"],
                r["reason"],
                r["resolved_at"].isoformat() if r["resolved_at"] else "",
                r["resolved_by"],
                r["resolution_notes"],
            ]
            for r in qd["rows"]
        ]
        or [["—", "", "", "", "", ""]],
    )

    # ── Dérivations ──
    ws = wb.create_sheet("Dérivations")
    ws["A1"] = "Résumé des dérivations (CDC §8.5)"
    ws["A1"].font = _TITLE_FONT
    d = data["derivations"]
    row = _write_table(
        ws,
        3,
        ["Dérivation", "Nombre"],
        [
            ["Produits indexés cuivre", d["copper_indexed"]],
            ["factory_code renseigné", d["factory_code_filled"]],
            ["parent_reference renseigné", d["parent_reference_filled"]],
        ],
    )
    _write_table(
        ws,
        row,
        ["Unité de base", "Nombre"],
        [[b["value"], b["count"]] for b in d["base_unit"]] or [["—", 0]],
    )

    # ── Simulation ──
    ws = wb.create_sheet("Simulation")
    ws["A1"] = "Statistiques simulation initiale"
    ws["A1"].font = _TITLE_FONT
    sim = data["simulations"]
    if sim["available"]:
        ws["A2"] = (
            f"{sim['finalized_line_count']} ligne(s) finalisée(s) · {sim['anomalies']} anomalie(s)"
        )
        _write_table(
            ws,
            4,
            ["Gamme", "PV moyen (€)", "Nb lignes"],
            [
                [
                    r["range"],
                    round(r["avg_pv_eur"], 2) if r["avg_pv_eur"] is not None else "",
                    r["count"],
                ]
                for r in sim["pv_by_range"]
            ],
        )
    else:
        ws["A2"] = "Aucune simulation finalisée — pas de statistiques de prix."

    # Reasonable default column widths on every sheet.
    for sheet in wb.worksheets:
        for col_cells in sheet.columns:
            width = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
            sheet.column_dimensions[col_cells[0].column_letter].width = min(max(width + 2, 12), 60)

    return wb


# ── Email body ───────────────────────────────────────────────────────────────


def render_email_body(data: dict) -> str:
    """Short FR plain-text summary for the cross-validation email."""
    t = data["totals"]
    qd = data["quarantine"]
    sources = ", ".join(f"{s['value']}={s['count']}" for s in data["products_by_source"]) or "—"
    lines = [
        "Rapport de migration Syskern — synthèse",
        f"Généré le {data['generated_at']}",
        "",
        f"Produits        : {t['products']} (par source : {sources})",
        f"Fournisseurs    : {t['suppliers']}",
        f"Clients         : {t['clients']}",
        f"Attributs       : {t['attributes']}",
        f"Quarantaine     : {qd['total']} ligne(s) — {qd['unresolved']} à traiter",
        f"  par raison    : {qd['by_reason'] or '—'}",
        f"Indexés cuivre  : {data['derivations']['copper_indexed']}",
    ]
    sim = data["simulations"]
    if sim["available"]:
        lines.append(
            f"Simulation      : {sim['finalized_line_count']} ligne(s) finalisée(s), "
            f"{len(sim['pv_by_range'])} gamme(s), {sim['anomalies']} anomalie(s)"
        )
    else:
        lines.append("Simulation      : aucune simulation finalisée")
    lines += [
        "",
        "Détail complet dans le fichier Excel joint (onglets Synthèse, Fournisseurs, "
        "Attributs, Quarantaine, Dérivations, Simulation).",
        "Merci de valider avant mise en production.",
    ]
    return "\n".join(lines)


# ── Orchestration ────────────────────────────────────────────────────────────


def generate_report(
    *,
    output_dir: str | Path,
    when: datetime | None = None,
    state_file: str | Path | None = None,
) -> tuple[Path, str]:
    """Build the data, write the Excel file, and return ``(path, email_body)``."""
    stamp = when or datetime.now()
    data = build_report_data(state_file=state_file, when=stamp)
    wb = build_workbook(data)

    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    path = out / f"migration_report_{stamp:%Y-%m-%d_%H-%M-%S}.xlsx"
    wb.save(path)
    logger.info("Migration report written to %s", path)
    return path, render_email_body(data)
