"""Tests for MirsanLoader — no-DB unit tests only (multi-sheet row normalisation).

DB integration tests (ProductSupplier upsert, quarantine) require Docker Postgres
and follow the same pattern as test_loader_po_fournisseurs.py.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from apps.data_migration.loaders.loader_po_mirsan import (
    _FACTORY_CODE,
    _SHEET_PROFILES,
    MirsanLoader,
    _is_valid_ukn_code,
    _read_sheet_as_df,
    _row_cell,
)

FIXTURE_DIR = Path(__file__).parent / "fixtures"
GEN_DIR = FIXTURE_DIR / "_generated"  # runtime-generated fixtures (gitignored)
MIRSAN_FIXTURE = GEN_DIR / "mirsan_sample.xlsx"

# ─── Fixture generation ───────────────────────────────────────────────────────


def _generate_mirsan_fixture(path: Path) -> None:
    wb = openpyxl.Workbook()

    # START CABINETS — header_row=5 (0-based), UKN in col 1, price in col 16
    ws = wb.active
    ws.title = "START CABINETS"
    for _ in range(5):
        ws.append([None] * 25)
    ws.append(
        [
            "Mirsan Product Code",
            "ukn products codes ",
            "ukn number codes",
            "Dimensions",
            "Designation ",
        ]
        + ["Unit Price"] * 11
        + [
            "SYMEA DDP Prices",
            "SYMEA DDP Prices",
            None,
            "SYSKERN DDP Prices",
            "SYSKERN DDP Prices",
            None,
            "Ex Works Eskisehir",
        ]
    )
    ws.append(
        [
            "MR.HD.GTN22U66.01_STRS",
            "KRS2266SGR",
            "U1402001",
            "22RU 600mmx600mm",
            "Single glass door - Rear panel",
        ]
        + [None] * 11
        + [124.31, None, None, 132.25, None, None, None]
    )
    ws.append(
        [
            "MR.HD.GTN24U66.01_STRS",
            "KRS2466SGR",
            "U1402002",
            "24RU 600mmx600mm",
            "Single glass door - Rear panel",
        ]
        + [None] * 11
        + [133.0, None, None, 141.0, None, None, 158.0]
    )
    # Blank row (section break)
    ws.append([None] * 25)
    ws.append(
        [
            "MR.HD.GTN42U66.01_STRS",
            "KRS4266SGR",
            "U1402005",
            "42RU 600mmx600mm",
            "Single glass door - Rear panel",
        ]
        + [None] * 11
        + [184.0, None, None, 196.0, None, None, 205.0]
    )

    # GRID CABINETS — header_row=4 (0-based), price in col 13
    ws2 = wb.create_sheet("GRID CABINETS")
    for _ in range(4):
        ws2.append([None] * 20)
    ws2.append(
        [
            "Mirsan Product Code",
            "ukn products codes ",
            "ukn number codes",
            "Dimensions",
            "Door Configurations",
        ]
        + ["Unit Price"] * 4
        + ["TG", "Unit Price", "SYMEA DDP Prices", None, "SYSKERN DDP Prices"]
    )
    ws2.append(
        [
            "MR.HD.GTN22U66G.01_GRID",
            "KRG2266SGR",
            "U1403001",
            "22RU 600mmx600mm",
            "Single Front Glass Door - Full rear",
            193.1,
            None,
            None,
            None,
            161,
            None,
            151.4,
            None,
            161.1,
        ]
    )
    ws2.append(
        [
            "MR.HD.GTN24U66G.01_GRID",
            "KRG2466SGR",
            "U1403002",
            "24RU 600mmx600mm",
            "Single Front Glass Door - Full rear",
            201.7,
            None,
            None,
            None,
            172,
            None,
            165.0,
            None,
            175.5,
        ]
    )

    # RACKS & OPEN RACKS — header_row=4 (0-based), UKN in col 1, price in col 13
    ws3 = wb.create_sheet("RACKS & OPEN RACKS")
    for _ in range(4):
        ws3.append([None] * 16)
    ws3.append(
        ["Product Description", "ukn products codes ", "ukn number codes", "Mirsan Code"]
        + ["SYMEA PRICES"] * 5
        + ["TG"]
        + ["SYMEA PRICES"] * 3
        + [None, "SYSKERN DDP Prices"]
    )
    ws3.append(
        [
            "691707IP55.2",
            "KWI07600",
            "-",
            "MR.IP55W07U60.03",
            166.93,
            136.85,
            None,
            None,
            "$120,00",
            120.0,
            None,
            102.92,
            None,
            109.49,
        ]
    )
    ws3.append(
        [
            "691709IP55.2",
            "KWI09600",
            "-",
            "MR.IP55W09U60.03",
            175.34,
            152.95,
            None,
            None,
            "$130,00",
            130.0,
            None,
            111.49,
            None,
            118.61,
        ]
    )

    # ACCESSORIES 19 — header_row=3 (0-based), UKN in col 2, price in col 9
    ws4 = wb.create_sheet("ACCESSORIES 19")
    for _ in range(3):
        ws4.append([None] * 12)
    ws4.append(
        [
            "MIRSAN CODE",
            "DESIGNATION / TYPE",
            "ukn products codes ",
            "ukn number codes",
            "Product Description",
        ]
        + ["EX-WORKS PRICES ESKISEHIR"] * 4
        + ["SYMEA DDP Prices"]
    )
    ws4.append(
        [
            "MR.FAN2AT.XX",
            "FAN BLOCK 2V",
            "KFAN2V",
            "ACC001",
            "2 fans analog thermostat module",
            30.25,
            30.25,
            25.0,
            25.0,
            21.44,
        ]
    )
    ws4.append(
        [
            "MR.FAN4AT.XX",
            "FAN BLOCK 4V",
            "KFAN4V",
            "ACC002",
            "4 fans analog thermostat module",
            43.45,
            43.45,
            40.0,
            40.0,
            34.31,
        ]
    )

    wb.save(path)


@pytest.fixture(scope="module", autouse=True)
def create_mirsan_fixture() -> None:
    GEN_DIR.mkdir(parents=True, exist_ok=True)
    _generate_mirsan_fixture(MIRSAN_FIXTURE)


# ─── Unit tests (no DB) ───────────────────────────────────────────────────────


class TestIsValidUKNCode:
    def test_valid_krs(self) -> None:
        assert _is_valid_ukn_code("KRS2266SGR") is True

    def test_valid_krg(self) -> None:
        assert _is_valid_ukn_code("KRG4266SGR") is True

    def test_valid_kwi(self) -> None:
        assert _is_valid_ukn_code("KWI07600") is True

    def test_valid_kfan(self) -> None:
        assert _is_valid_ukn_code("KFAN2V") is True

    def test_invalid_none(self) -> None:
        assert _is_valid_ukn_code(None) is False

    def test_invalid_header(self) -> None:
        assert _is_valid_ukn_code("Mirsan Product Code") is False

    def test_invalid_dimension(self) -> None:
        assert _is_valid_ukn_code("22RU 600x600") is False

    def test_invalid_dash(self) -> None:
        assert _is_valid_ukn_code("-") is False


class TestReadSheetAsDf:
    def test_start_cabinets_row_count(self) -> None:
        import openpyxl as opxl

        wb = opxl.load_workbook(str(MIRSAN_FIXTURE), data_only=True, read_only=True)
        profile = next(p for p in _SHEET_PROFILES if p.sheet_name == "START CABINETS")
        df = _read_sheet_as_df(wb, profile)
        assert len(df) == 3  # KRS2266SGR, KRS2466SGR, KRS4266SGR

    def test_racks_sheet_price_syskern_ddp(self) -> None:
        import openpyxl as opxl

        wb = opxl.load_workbook(str(MIRSAN_FIXTURE), data_only=True, read_only=True)
        profile = next(p for p in _SHEET_PROFILES if p.sheet_name == "RACKS & OPEN RACKS")
        df = _read_sheet_as_df(wb, profile)
        # price_col=13 → SYSKERN DDP = 109.49 EUR (not the "$120,00" TG col)
        price_val = df.loc[df["sku_code"] == "KWI07600", "price_eur"].iloc[0]
        assert price_val is not None
        assert float(price_val) == pytest.approx(109.49)

    def test_accessories_sku_col_2(self) -> None:
        import openpyxl as opxl

        wb = opxl.load_workbook(str(MIRSAN_FIXTURE), data_only=True, read_only=True)
        profile = next(p for p in _SHEET_PROFILES if p.sheet_name == "ACCESSORIES 19")
        df = _read_sheet_as_df(wb, profile)
        assert "KFAN2V" in df["sku_code"].values
        assert "KFAN4V" in df["sku_code"].values

    def test_sheet_not_found_returns_empty(self) -> None:
        import openpyxl as opxl

        from apps.data_migration.loaders.loader_po_mirsan import _SheetProfile

        wb = opxl.load_workbook(str(MIRSAN_FIXTURE), data_only=True, read_only=True)
        missing = _SheetProfile(
            sheet_name="DOES NOT EXIST",
            header_row=0,
            sku_col=0,
            item_code_col=1,
            desc_col=2,
            mirsan_col=3,
            dim_col=4,
            price_col=5,
        )
        df = _read_sheet_as_df(wb, missing)
        assert df.empty


class TestRowCellBounds:
    """Regression: col index -1 means absent, not Python row[-1]."""

    def test_negative_index_returns_none(self) -> None:
        row = ("sku", "last")
        assert _row_cell(row, -1, len(row)) is None

    def test_valid_index(self) -> None:
        row = ("a", "b", "c")
        assert _row_cell(row, 1, len(row)) == "b"

    def test_out_of_range_returns_none(self) -> None:
        row = ("a", "b")
        assert _row_cell(row, 10, len(row)) is None


class TestMirsanNormalizeRow:
    def _make_raw(self, **kwargs: object) -> pd.Series:
        defaults = {
            "sku_code": "KRS2266SGR",
            "item_code": "U1402001",
            "description": "Single glass door - Rear panel",
            "mirsan_code": "MR.HD.GTN22U66.01_STRS",
            "dimensions": "22RU 600mmx600mm",
            "price_eur": "132.25",
            "__sheet__": "START CABINETS",
        }
        defaults.update(kwargs)
        return pd.Series(defaults)

    def test_sku_code(self) -> None:
        loader = MirsanLoader()
        row = loader.normalize_row(self._make_raw())
        assert row.data["sku_code"] == "KRS2266SGR"

    def test_price_decimal(self) -> None:
        loader = MirsanLoader()
        row = loader.normalize_row(self._make_raw(price_eur="132.25"))
        assert row.data["price_eur"] == Decimal("132.25")

    def test_desc_with_dimensions_combined(self) -> None:
        loader = MirsanLoader()
        row = loader.normalize_row(
            self._make_raw(
                description="Single glass door",
                dimensions="22RU 600x600",
            )
        )
        assert row.data["description"] == "Single glass door — 22RU 600x600"

    def test_desc_without_dimensions(self) -> None:
        loader = MirsanLoader()
        row = loader.normalize_row(self._make_raw(description="Fan Block 2V", dimensions=None))
        assert row.data["description"] == "Fan Block 2V"

    def test_mirsan_code_in_data(self) -> None:
        loader = MirsanLoader()
        row = loader.normalize_row(self._make_raw())
        assert row.data["mirsan_code"] == "MR.HD.GTN22U66.01_STRS"

    def test_null_price_is_none(self) -> None:
        loader = MirsanLoader()
        row = loader.normalize_row(self._make_raw(price_eur=None))
        assert row.data["price_eur"] is None


class TestMirsanMatchHint:
    def test_sku_code_in_hint(self) -> None:
        loader = MirsanLoader()
        raw = pd.Series(
            {
                "sku_code": "KRS2266SGR",
                "price_eur": "132",
                "item_code": None,
                "description": "desc",
                "mirsan_code": None,
                "dimensions": None,
                "__sheet__": "START CABINETS",
            }
        )
        hint = loader.build_match_hint(loader.normalize_row(raw))
        assert hint.sku_code == "KRS2266SGR"
        assert hint.factory_code == _FACTORY_CODE
