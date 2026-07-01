"""Tests for the final post-migration report (CDC §8.8)."""

from __future__ import annotations

from decimal import Decimal

import pytest
from openpyxl import load_workbook

from apps.attributes.models import AttributeRegistry
from apps.clients.models import Client
from apps.data_migration.models import MigrationUnmatched, UnmatchedReason
from apps.data_migration.report import (
    build_report_data,
    build_workbook,
    generate_report,
    render_email_body,
)
from apps.products.models import Product, ProductSupplier
from apps.simulations.models import Simulation, SimulationLine

pytestmark = pytest.mark.django_db


@pytest.fixture()
def seeded():
    odoo = Product.objects.create(
        sku_code="KCFF6A4PZHDBL5-21",
        name="Câble cat7",
        range="Catégorie 7",
        migration_source="odoo",
        is_copper_indexed=True,
        copper_weight_kg_per_unit=Decimal("18"),
        factory_code="21",
        parent_reference="KCFF6A4PZHDBL5",
    )
    Product.objects.create(sku_code="RACK-1", name="Rack", migration_source="database_internal")
    ProductSupplier.objects.create(
        product=odoo, supplier_name="Symea", is_active=True, po_base_price=Decimal("2350")
    )
    AttributeRegistry.objects.create(
        code="shielding", label={"fr": "Blindage"}, category="technical", data_type="text"
    )
    Client.objects.create(name="Acme")
    MigrationUnmatched.objects.create(
        source_file="PO.xlsx",
        source_row_number=5,
        raw_data={"x": 1},
        reason=UnmatchedReason.NO_MATCH,
    )

    # Draft → add line → finalize: the guard trigger blocks line inserts on a
    # finalized parent, so the line must exist before the simulation is frozen.
    sim = Simulation.objects.create(label="Tarif Q3", simulation_type="tariff", status="draft")
    SimulationLine.objects.create(
        simulation=sim, product=odoo, pv_eur=Decimal("487.70"), status="ok"
    )
    Simulation.objects.filter(pk=sim.pk).update(status="finalized")
    return odoo


def test_build_report_data_aggregates(seeded):
    data = build_report_data()

    assert data["totals"]["products"] == 2
    assert data["totals"]["suppliers"] == 1
    assert data["totals"]["clients"] == 1
    # AttributeRegistry has a seeded baseline (CDC §3.3) + the one we created;
    # assert the report counts the live total, not a hardcoded number.
    assert data["totals"]["attributes"] == AttributeRegistry.objects.count() >= 1
    assert data["totals"]["quarantine"] == 1

    sources = {row["value"]: row["count"] for row in data["products_by_source"]}
    assert sources == {"odoo": 1, "database_internal": 1}

    assert data["quarantine"]["unresolved"] == 1
    assert data["quarantine"]["by_reason"] == {"no_match": 1}

    assert data["derivations"]["copper_indexed"] == 1
    assert data["derivations"]["factory_code_filled"] == 1

    assert data["simulations"]["available"] is True
    pv = {r["range"]: r["avg_pv_eur"] for r in data["simulations"]["pv_by_range"]}
    assert pv["Catégorie 7"] == pytest.approx(487.70)


def test_build_workbook_has_all_tabs(seeded):
    wb = build_workbook(build_report_data())
    assert wb.sheetnames == [
        "Synthèse",
        "Fournisseurs",
        "Attributs",
        "Quarantaine",
        "Dérivations",
        "Simulation",
    ]


def test_render_email_body_has_headline_figures(seeded):
    body = render_email_body(build_report_data())
    assert "Rapport de migration" in body
    assert "Produits" in body
    assert "Quarantaine" in body


def test_generate_report_writes_xlsx(seeded, tmp_path):
    path, body = generate_report(output_dir=tmp_path)
    assert path.exists()
    assert path.name.startswith("migration_report_")
    assert path.suffix == ".xlsx"
    # File is a valid workbook with the expected tabs.
    wb = load_workbook(path)
    assert "Synthèse" in wb.sheetnames
    assert body  # non-empty email summary


def test_report_without_simulation_is_graceful(db):
    Product.objects.create(sku_code="X-1", name="X", migration_source="manual")
    data = build_report_data()
    assert data["simulations"]["available"] is False
    wb = build_workbook(data)  # must not raise
    assert "Simulation" in wb.sheetnames
