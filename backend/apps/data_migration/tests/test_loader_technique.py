"""Tests for TechniqueLoader (CDC §8.4 — Étape 2, Phase 3).

Structure
---------
- ``TestTechCleanHelper``      : unit tests on _clean() helper (no DB)
- ``TestTechNormalizeRow``     : unit tests on normalize_row() (no DB)
- ``TestTechDedup``            : unit tests on dedup_key() (no DB)
- ``TestTechMatchHint``        : unit tests on build_match_hint() (no DB)
- ``TestTechFixturePrepare``   : integration tests with the generated xlsx fixture (no DB)
- ``TestTechLoaderReport``     : @pytest.mark.django_db — full run against DB
- ``TestTechProductEnrichment``: @pytest.mark.django_db — Product field enrichment
- ``TestTechEAV``              : @pytest.mark.django_db — EAV upsert behaviour
- ``TestTechDryRun``           : @pytest.mark.django_db — dry_run mode
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from apps.data_migration.loaders.exceptions import HeaderValidationError
from apps.data_migration.loaders.loader_technique import (
    _SHEET_PROFILES,
    TechniqueLoader,
    _clean,
    _prepare_sheet_df,
)
from apps.data_migration.loaders.types import LoaderConfig
from apps.data_migration.models import MigrationUnmatched, UnmatchedReason
from apps.products.models import MigrationSource, Product

FIXTURE_DIR = Path(__file__).parent / "fixtures"
GEN_DIR = FIXTURE_DIR / "_generated"  # runtime-generated fixtures (gitignored)
TECH_FIXTURE = GEN_DIR / "technique_sample.xlsx"

# ─── Fixture generation ───────────────────────────────────────────────────────

# Column order for UKN sheet (matches real file after pandas rename)
_UKN_HEADERS = [
    "Range",
    "Sub-Range",
    "Type",
    "AWG",
    "Internal Code",
    "Unikkern Items code",
    "Description En",
    "OD (mm)",
    "lab",
    "CPR level",
    "CPR updated date",
    "Report no.",
    "Cu (kg/km)",
    "Origin",
    "Per Product Net Content",
    "Per Product Net Content (Unit)",
    "Global Trade Item Number (GTIN)",
    "Individual Qty",
    "Individual carton/bag size(mm)",
    "Qty per inner carton",
    "Inner carton size(cm)",
    "Qty per export carton",
    "Export carton size(mm)",
    "QTY/Pallet",
    "Loading Way /Pallet",
    "Pallet Size cm",
    None,
    "Sticker ENG",
    "Sticker FR",
    "UID Code",
    "DoP",
    "Marking",
]

_NX_HEADERS = [
    "Range",
    "Sub-Range",
    "Type",
    "AWG",
    "Unikkern Items code",
    "Description En",
    "CPR",
    "Cu (kg/km)",
    "Origin",
    "Per Product Net Content",
    "Per Product Net Content (Unit)",
    "Global Trade Item Number (GTIN)",
    "Individual Qty",
    "Individual carton/bag size(mm)",
    "Qty per inner carton",
    "Inner carton size(cm)",
    "Qty per export carton",
    "Export carton size(mm)",
    "QTY/Pallet",
    "Loading Way /Pallet",
    "Pallet Size cm",
    None,
    "Sticker ENG",
    "Sticker FR",
    "UID Code",
    "DoP",
    "Marking",
]

_OR_HEADERS = [
    "Range",
    "Sub-Range",
    "Type",
    "AWG",
    "Unikkern Items code",
    "Description En",
    "CPR",
    "Origin",
    "Cu (kg/km)",
    "Individual Packing",
    "Per Product Net Content",
    "Per Product Net Content (Unit)",
    "Global Trade Item Number (GTIN)",
    "Individual Qty",
    "Individual carton/bag size",
    "Qty per inner carton",
    "Inner carton size mm",
    "Qty per export carton",
    "Export carton size mm",
    "QTY/Pallet",
    "Pallet Size cm",
    None,
    "Sticker ENG",
    "Sticker FR",
    None,
]


def _ukn_row(**kwargs: object) -> list:
    """Build a UKN data row with sensible defaults."""
    defaults = {
        "Range": "SOLID CABLE",
        "Sub-Range": "SOLID CABLE CAT6",
        "Type": "U/UTP",
        "AWG": "23",
        "Internal Code": "KTEST001-21",
        "Unikkern Items code": "KTEST001",
        "Description En": "CAT6 TEST CABLE 500M",
        "OD (mm)": "5.80 +/- 0.3",
        "lab": "SGS - 9999",
        "CPR level": "Dca-s2, d1, a1",
        "CPR updated date": None,
        "Report no.": None,
        "Cu (kg/km)": "16.0",
        "Origin": "China",
        "Per Product Net Content": "62",
        "Per Product Net Content (Unit)": "Kilogram (kg)",
        "Global Trade Item Number (GTIN)": "4897108899999",
        "Individual Qty": None,
        "Individual carton/bag size(mm)": None,
        "Qty per inner carton": None,
        "Inner carton size(cm)": None,
        "Qty per export carton": None,
        "Export carton size(mm)": None,
        "QTY/Pallet": "12",
        "Loading Way /Pallet": None,
        "Pallet Size cm": None,
        None: None,
        "Sticker ENG": "CAT6 TEST CABLE 500M ENG",
        "Sticker FR": "CÂBLE CAT6 TEST 500M FR",
        "UID Code": "UKN-TEST001DLH",
        "DoP": "UKN-L07TEST-XX",
        "Marking": None,
    }
    defaults.update(kwargs)
    return [defaults.get(h) for h in _UKN_HEADERS]


def _nx_row(**kwargs: object) -> list:
    defaults = {
        "Range": "TELEPHONE CABLE",
        "Sub-Range": "TELEPHONE CABLE SYT1",
        "Type": None,
        "AWG": None,
        "Unikkern Items code": "NKTEST001",
        "Description En": "NEXKERN TEST CABLE",
        "CPR": None,
        "Cu (kg/km)": None,
        "Origin": "TURKEY",
        "Per Product Net Content": "19.5",
        "Per Product Net Content (Unit)": "Kilogram (kg) [KGM]",
        "Global Trade Item Number (GTIN)": "4897108811111",
        "Individual Qty": None,
        "Individual carton/bag size(mm)": None,
        "Qty per inner carton": None,
        "Inner carton size(cm)": None,
        "Qty per export carton": None,
        "Export carton size(mm)": None,
        "QTY/Pallet": None,
        "Loading Way /Pallet": None,
        "Pallet Size cm": None,
        None: None,
        "Sticker ENG": "NEXKERN TEST ENG",
        "Sticker FR": "NEXKERN TEST FR",
        "UID Code": "#REF!",  # Excel error — should be ignored
        "DoP": "XKN-TEST-XX-0001",
        "Marking": None,
    }
    defaults.update(kwargs)
    return [defaults.get(h) for h in _NX_HEADERS]


def generate_technique_fixture(path: Path) -> None:
    wb = openpyxl.Workbook()

    # ── UKN sheet ─────────────────────────────────────────────────────────────
    ws_ukn = wb.active
    ws_ukn.title = "GTIN code & packing details"
    ws_ukn.append(_UKN_HEADERS)

    # Scenario 1: full match + enrichment + 3 EAV
    ws_ukn.append(_ukn_row())

    # Scenario 2a: China row of a duplicate SKU (most complete)
    ws_ukn.append(
        _ukn_row(
            **{
                "Internal Code": "KTEST002-21",
                "Unikkern Items code": "KTEST002",
                "GTIN": "4897108888881",
                "Origin": "China",
                "UID Code": "UKN-TEST002DLH",
            }
        )
    )
    # Scenario 2b: Turkey row of same SKU (less complete — OD and CPR missing)
    ws_ukn.append(
        _ukn_row(
            **{
                "Internal Code": "KTEST002-E02",
                "Unikkern Items code": "KTEST002",
                "Global Trade Item Number (GTIN)": "4897108888881",
                "Origin": "Turkey",
                "OD (mm)": None,
                "CPR level": None,
                "UID Code": "#REF!",
                "Cu (kg/km)": None,
                "QTY/Pallet": None,
            }
        )
    )

    # Scenario 3: no SKU → NO_SKU quarantine
    ws_ukn.append(_ukn_row(**{"Unikkern Items code": None, "Internal Code": None}))

    # Scenario 4: SKU not in DB → NO_MATCH quarantine
    ws_ukn.append(
        _ukn_row(
            **{
                "Internal Code": "UNKNOWNSKUXX-21",
                "Unikkern Items code": "UNKNOWNSKUXX",
                "UID Code": "#REF!",
            }
        )
    )

    # ── NEXKERN sheet ─────────────────────────────────────────────────────────
    ws_nx = wb.create_sheet("NEXKERN GTIN code & packing")
    ws_nx.append(_NX_HEADERS)

    # Scenario 5: NEXKERN with UID #REF! — uid_code ignored, rest enriched
    ws_nx.append(_nx_row())

    # ── ORSEAN sheet ──────────────────────────────────────────────────────────
    ws_or = wb.create_sheet("ORSEAN GTIN code & packing")
    ws_or.append(_OR_HEADERS)
    # Minimal row — ORSEAN has no UID/DoP columns
    ws_or.append(
        [
            "SOLID CABLE",
            "TV",
            "F/FTP",
            "23",
            "OEG3TEST5",
            "ORSEAN TEST CABLE",
            "Dca",
            "China",
            "22",
            "500m drum",
            "56",
            "Kilogram (kg)",
            "4897108884031",
            None,
            None,
            None,
            None,
            None,
            None,
            "8",
            None,
            None,
            None,
            None,
            None,
        ]
    )

    wb.save(path)


@pytest.fixture(scope="module", autouse=True)
def create_technique_fixture() -> None:
    GEN_DIR.mkdir(parents=True, exist_ok=True)
    generate_technique_fixture(TECH_FIXTURE)


def default_config(**kw: object) -> LoaderConfig:
    return LoaderConfig(
        file_path=str(TECH_FIXTURE),
        sheet_name="GTIN code & packing details",
        header_row=0,
        batch_size=500,
        dry_run=bool(kw.get("dry_run", False)),
    )


def make_product(sku_code: str, **kwargs: object) -> Product:
    return Product.objects.create(
        sku_code=sku_code,
        name=f"Test {sku_code}",
        migration_source=MigrationSource.MANUAL,
        **kwargs,
    )


# ─── _clean helper ────────────────────────────────────────────────────────────


class TestTechCleanHelper:
    def test_none_returns_none(self) -> None:
        assert _clean(None) is None

    def test_excel_ref_error(self) -> None:
        assert _clean("#REF!") is None

    def test_na_error(self) -> None:
        assert _clean("#N/A") is None

    def test_dash(self) -> None:
        assert _clean("-") is None

    def test_valid_string(self) -> None:
        assert _clean("UKN-C6TEST") == "UKN-C6TEST"

    def test_strips_whitespace(self) -> None:
        assert _clean("  Dca-s2 ") == "Dca-s2"


# ─── normalize_row ────────────────────────────────────────────────────────────


class TestTechNormalizeRow:
    def _make_raw(self, **kwargs: object) -> pd.Series:
        defaults = {
            "sku_code": "KTEST001",
            "internal_code": "KTEST001-21",
            "__brand__": "ukn",
            "cpr_level": "Dca-s2, d1, a1",
            "od_mm": "5.80 +/- 0.3",
            "cu_kg_km": "16.0",
            "net_content": "62",
            "net_unit": "Kilogram (kg)",
            "gtin": "4897108899999",
            "pallet_qty": "12",
            "uid_code": "UKN-TEST001DLH",
            "dop_number": "UKN-L07TEST-XX",
            "sticker_en": "CAT6 TEST 500M ENG",
            "sticker_fr": "CÂBLE CAT6 TEST 500M FR",
        }
        defaults.update(kwargs)
        return pd.Series(defaults)

    def test_sku_code(self) -> None:
        row = TechniqueLoader().normalize_row(self._make_raw())
        assert row.data["sku_code"] == "KTEST001"

    def test_copper_decimal(self) -> None:
        row = TechniqueLoader().normalize_row(self._make_raw())
        assert row.data["copper_kg_km"] == Decimal("16.0")

    def test_unit_weight_kg_when_unit_is_kg(self) -> None:
        row = TechniqueLoader().normalize_row(self._make_raw())
        assert row.data["unit_weight_kg"] == Decimal("62")

    def test_unit_weight_ignored_when_unit_not_kg(self) -> None:
        row = TechniqueLoader().normalize_row(self._make_raw(net_unit="Metre (m)"))
        assert row.data["unit_weight_kg"] is None

    def test_gtin_clean(self) -> None:
        row = TechniqueLoader().normalize_row(self._make_raw())
        assert row.data["gtin"] == "4897108899999"

    def test_gtin_na_is_none(self) -> None:
        row = TechniqueLoader().normalize_row(self._make_raw(gtin="#N/A"))
        assert row.data["gtin"] is None

    def test_uid_code_ref_is_none(self) -> None:
        row = TechniqueLoader().normalize_row(self._make_raw(uid_code="#REF!"))
        assert row.data["eav_uid_code"] is None

    def test_uid_code_valid(self) -> None:
        row = TechniqueLoader().normalize_row(self._make_raw())
        assert row.data["eav_uid_code"] == "UKN-TEST001DLH"

    def test_dop_ref_is_none(self) -> None:
        row = TechniqueLoader().normalize_row(self._make_raw(dop_number="#REF!"))
        assert row.data["dop_number"] is None

    def test_factory_code_extracted(self) -> None:
        row = TechniqueLoader().normalize_row(self._make_raw())
        assert row.data["factory_code"] == "21"

    def test_factory_code_turkey(self) -> None:
        row = TechniqueLoader().normalize_row(self._make_raw(internal_code="KTEST001-E02"))
        assert row.data["factory_code"] == "E02"


# ─── dedup_key ────────────────────────────────────────────────────────────────


class TestTechDedup:
    def _norm(self, sku: str, brand: str) -> object:
        raw = pd.Series(
            {
                "sku_code": sku,
                "__brand__": brand,
                "cu_kg_km": "1",
                "net_content": "1",
                "net_unit": "Kilogram (kg)",
                "gtin": None,
                "pallet_qty": None,
                "uid_code": None,
                "dop_number": None,
                "sticker_en": None,
                "sticker_fr": None,
                "cpr_level": None,
                "od_mm": None,
                "internal_code": None,
            }
        )
        return TechniqueLoader().normalize_row(raw)

    def test_ukn_brand_returns_sku(self) -> None:
        n = self._norm("KTEST001", "ukn")
        assert TechniqueLoader().dedup_key(n) == "KTEST001"

    def test_nexkern_returns_none(self) -> None:
        n = self._norm("NKTEST001", "nexkern")
        assert TechniqueLoader().dedup_key(n) is None

    def test_orsean_returns_none(self) -> None:
        n = self._norm("OETEST001", "orsean")
        assert TechniqueLoader().dedup_key(n) is None


# ─── build_match_hint ─────────────────────────────────────────────────────────


class TestTechMatchHint:
    def _hint(self, sku: str, internal: str | None = None) -> object:
        raw = pd.Series(
            {
                "sku_code": sku,
                "__brand__": "ukn",
                "internal_code": internal,
                "cu_kg_km": "1",
                "net_content": "1",
                "net_unit": "Kilogram (kg)",
                "gtin": None,
                "pallet_qty": None,
                "uid_code": None,
                "dop_number": None,
                "sticker_en": None,
                "sticker_fr": None,
                "cpr_level": None,
                "od_mm": None,
            }
        )
        loader = TechniqueLoader()
        norm = loader.normalize_row(raw)
        return loader.build_match_hint(norm)

    def test_sku_code_in_hint(self) -> None:
        hint = self._hint("KTEST001", "KTEST001-21")
        assert hint.sku_code == "KTEST001"

    def test_parent_reference_equals_sku(self) -> None:
        hint = self._hint("KTEST001", "KTEST001-21")
        assert hint.parent_reference == hint.sku_code

    def test_factory_code_from_internal(self) -> None:
        hint = self._hint("KTEST001", "KTEST001-21")
        assert hint.factory_code == "21"


class TestTechHeaderValidation:
    """Per-sheet canonical headers must raise HeaderValidationError when absent."""

    def test_validate_raises_when_canonical_column_missing(self) -> None:
        loader = TechniqueLoader()
        profile = next(p for p in _SHEET_PROFILES if p.brand_tag == "ukn")
        loader._active_tech_profile = profile
        try:
            bad = pd.DataFrame({"sku_code": ["K1"]})
            with pytest.raises(HeaderValidationError) as exc:
                loader._validate_header(bad)
            assert "gtin" in exc.value.missing
        finally:
            loader._active_tech_profile = None


# ─── _prepare_sheet_df ────────────────────────────────────────────────────────


class TestTechFixturePrepare:
    def test_ukn_sheet_loaded(self) -> None:
        profile = next(p for p in _SHEET_PROFILES if p.brand_tag == "ukn")
        df = _prepare_sheet_df(str(TECH_FIXTURE), profile)
        assert "sku_code" in df.columns
        assert "__brand__" in df.columns
        assert (df["__brand__"] == "ukn").all()

    def test_ukn_rows_have_skus(self) -> None:
        profile = next(p for p in _SHEET_PROFILES if p.brand_tag == "ukn")
        df = _prepare_sheet_df(str(TECH_FIXTURE), profile)
        # 5 rows written: KTEST001, KTEST002 x2, empty SKU (dropped), UNKNOWNSKUXX
        assert len(df) == 4

    def test_nexkern_sheet_loaded(self) -> None:
        profile = next(p for p in _SHEET_PROFILES if p.brand_tag == "nexkern")
        df = _prepare_sheet_df(str(TECH_FIXTURE), profile)
        assert len(df) == 1
        assert df.iloc[0]["sku_code"] == "NKTEST001"

    def test_orsean_sheet_has_brand(self) -> None:
        profile = next(p for p in _SHEET_PROFILES if p.brand_tag == "orsean")
        df = _prepare_sheet_df(str(TECH_FIXTURE), profile)
        assert (df["__brand__"] == "orsean").all()


# ─── DB integration tests ─────────────────────────────────────────────────────


@pytest.mark.django_db(transaction=True)
class TestTechLoaderReport:
    def setup_method(self) -> None:
        make_product("KTEST001")
        make_product("KTEST002")
        make_product("NKTEST001")
        make_product("OEG3TEST5")
        # UNKNOWNSKUXX intentionally absent → NO_MATCH

    def test_total_rows(self) -> None:
        report = TechniqueLoader().run(default_config())
        # UKN: 4 rows (KTEST001 x1, KTEST002 x2 merged→1, UNKNOWNSKUXX x1) → 3 processable after dedup
        # + NEXKERN: 1 row, ORSEAN: 1 row
        # NO_SKU rows were dropped by _prepare_sheet_df so total processed = 3 UKN + 1 NX + 1 OR = 5
        assert report.rows_total == 5

    def test_matched_rows(self) -> None:
        report = TechniqueLoader().run(default_config())
        # KTEST001, KTEST002 (merged), NKTEST001, OEG3TEST5 = 4 matches
        assert report.rows_matched == 4

    def test_quarantine_no_match(self) -> None:
        report = TechniqueLoader().run(default_config())
        assert report.rows_unmatched.get(UnmatchedReason.NO_MATCH, 0) == 1

    def test_quarantine_entry_exists(self) -> None:
        TechniqueLoader().run(default_config())
        assert MigrationUnmatched.objects.filter(source_file="technique_sample.xlsx").exists()


@pytest.mark.django_db(transaction=True)
class TestTechProductEnrichment:
    def setup_method(self) -> None:
        make_product("KTEST001")
        make_product("KTEST002")
        make_product("NKTEST001")
        make_product("OEG3TEST5")

    def test_gtin_updated(self) -> None:
        TechniqueLoader().run(default_config())
        p = Product.objects.get(sku_code="KTEST001")
        assert p.gtin == "4897108899999"

    def test_copper_weight_set(self) -> None:
        TechniqueLoader().run(default_config())
        p = Product.objects.get(sku_code="KTEST001")
        assert p.copper_weight_kg_per_unit == Decimal("16.0")

    def test_is_copper_indexed_true(self) -> None:
        TechniqueLoader().run(default_config())
        p = Product.objects.get(sku_code="KTEST001")
        assert p.is_copper_indexed is True

    def test_unit_weight_kg(self) -> None:
        TechniqueLoader().run(default_config())
        p = Product.objects.get(sku_code="KTEST001")
        assert p.unit_weight_kg == Decimal("62")

    def test_pallet_qty(self) -> None:
        TechniqueLoader().run(default_config())
        p = Product.objects.get(sku_code="KTEST001")
        assert p.pallet_qty == 12

    def test_dop_number(self) -> None:
        TechniqueLoader().run(default_config())
        p = Product.objects.get(sku_code="KTEST001")
        assert p.dop_number == "UKN-L07TEST-XX"

    def test_sticker_en_description(self) -> None:
        TechniqueLoader().run(default_config())
        p = Product.objects.get(sku_code="KTEST001")
        assert p.description_marketing.get("en") == "CAT6 TEST CABLE 500M ENG"

    def test_sticker_fr_description(self) -> None:
        TechniqueLoader().run(default_config())
        p = Product.objects.get(sku_code="KTEST001")
        assert p.description_marketing.get("fr") == "CÂBLE CAT6 TEST 500M FR"

    def test_description_not_overwritten_if_exists(self) -> None:
        Product.objects.filter(sku_code="KTEST001").update(
            description_marketing={"en": "Existing English", "fr": "Existing French"}
        )
        TechniqueLoader().run(default_config())
        p = Product.objects.get(sku_code="KTEST001")
        assert p.description_marketing["en"] == "Existing English"

    def test_ktest002_merged_from_china_turkey(self) -> None:
        TechniqueLoader().run(default_config())
        p = Product.objects.get(sku_code="KTEST002")
        # China row had Cu=16, Turkey had Cu=None — merge keeps China's value
        assert p.copper_weight_kg_per_unit == Decimal("16.0")

    def test_nexkern_gtin_enriched(self) -> None:
        TechniqueLoader().run(default_config())
        p = Product.objects.get(sku_code="NKTEST001")
        assert p.gtin == "4897108811111"


@pytest.mark.django_db(transaction=True)
class TestTechEAV:
    def setup_method(self) -> None:
        make_product("KTEST001")
        make_product("KTEST002")
        make_product("NKTEST001")
        make_product("OEG3TEST5")

    def test_cpr_level_eav_created(self) -> None:
        from apps.attributes.models import AttributeRegistry, ProductAttributeValue

        TechniqueLoader().run(default_config())
        p = Product.objects.get(sku_code="KTEST001")
        attr = AttributeRegistry.objects.get(code="cpr_level")
        pav = ProductAttributeValue.objects.get(product=p, attribute=attr)
        assert pav.value == "Dca-s2, d1, a1"

    def test_od_mm_eav_created(self) -> None:
        from apps.attributes.models import AttributeRegistry, ProductAttributeValue

        TechniqueLoader().run(default_config())
        p = Product.objects.get(sku_code="KTEST001")
        attr = AttributeRegistry.objects.get(code="od_mm")
        pav = ProductAttributeValue.objects.get(product=p, attribute=attr)
        assert pav.value == "5.80 +/- 0.3"

    def test_uid_code_eav_created(self) -> None:
        from apps.attributes.models import AttributeRegistry, ProductAttributeValue

        TechniqueLoader().run(default_config())
        p = Product.objects.get(sku_code="KTEST001")
        attr = AttributeRegistry.objects.get(code="uid_code")
        pav = ProductAttributeValue.objects.get(product=p, attribute=attr)
        assert pav.value == "UKN-TEST001DLH"

    def test_uid_ref_error_not_written(self) -> None:
        from apps.attributes.models import AttributeRegistry, ProductAttributeValue

        TechniqueLoader().run(default_config())
        p = Product.objects.get(sku_code="NKTEST001")
        attr = AttributeRegistry.objects.get(code="uid_code")
        assert not ProductAttributeValue.objects.filter(product=p, attribute=attr).exists()

    def test_eav_idempotent(self) -> None:
        from apps.attributes.models import AttributeRegistry, ProductAttributeValue

        TechniqueLoader().run(default_config())
        TechniqueLoader().run(default_config())  # second run
        p = Product.objects.get(sku_code="KTEST001")
        attr = AttributeRegistry.objects.get(code="cpr_level")
        assert ProductAttributeValue.objects.filter(product=p, attribute=attr).count() == 1


@pytest.mark.django_db(transaction=True)
class TestTechDryRun:
    def setup_method(self) -> None:
        make_product("KTEST001")
        make_product("KTEST002")
        make_product("NKTEST001")
        make_product("OEG3TEST5")

    def test_dry_run_no_gtin_written(self) -> None:
        TechniqueLoader().run(default_config(dry_run=True))
        p = Product.objects.get(sku_code="KTEST001")
        assert p.gtin == ""

    def test_dry_run_no_eav_created(self) -> None:
        from apps.attributes.models import ProductAttributeValue

        TechniqueLoader().run(default_config(dry_run=True))
        assert ProductAttributeValue.objects.count() == 0

    def test_dry_run_report_populated(self) -> None:
        report = TechniqueLoader().run(default_config(dry_run=True))
        assert report.dry_run is True
        assert report.rows_matched == 4
